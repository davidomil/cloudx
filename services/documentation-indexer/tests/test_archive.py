from __future__ import annotations

import hashlib
import json
import shutil
import subprocess
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from importlib.metadata import version
from pathlib import Path

from fastapi.testclient import TestClient
from PIL import Image, ImageDraw
import pytest
from reportlab.pdfgen import canvas

import cloudx_documentation_indexer.archive as archive_module
from cloudx_documentation_indexer import DocumentationArchive, create_app
from cloudx_documentation_indexer.archive import (
    DENSE_ONLY_MIN_SCORE,
    EMBEDDING_DIM,
    ArchiveError,
    fetch_url_bytes,
    reciprocal_rank_fusion,
)
from cloudx_documentation_indexer.extraction import plausible_tables, table_span_text


def test_archive_ingests_searches_invalidates_and_remains_portable(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "fixtures"
    fixture_dir.mkdir()
    datasheet_pdf = fixture_dir / "reset-controller-datasheet.pdf"
    book_pdf = fixture_dir / "vector-search-handbook.pdf"
    readme = fixture_dir / "README.md"
    code = fixture_dir / "driver.c"
    make_pdf(
        datasheet_pdf,
        [
            "Reset Controller Datasheet",
            "BOOT_MODE register samples the boot pin during reset.",
            "RESET_N timing requires a ten millisecond low pulse.",
        ],
    )
    make_pdf(
        book_pdf,
        [
            "Vector Search Handbook",
            "Chapter 4 explains reciprocal rank fusion for hybrid retrieval.",
            "Portable indexes should store a manifest and all source snapshots.",
        ],
    )
    readme.write_text("Board README\nThe vendor board readme describes SWD debug wiring and boot jumper setup.\n", encoding="utf-8")
    code.write_text("void configure_boot_pin(void) { /* enable BOOT_MODE pull down before reset */ }\n", encoding="utf-8")

    archive = DocumentationArchive(tmp_path / "archive")
    archive.ingest_path(datasheet_pdf, source_type="datasheet", collection="board")
    archive.ingest_path(book_pdf, source_type="book", collection="books")
    archive.ingest_path(fixture_dir, collection="vendor", accept_generated_code_documentation=True)

    server = HtmlFixtureServer("<html><body><h1>Thermal Layout Note</h1><p>The website says copper pours improve regulator thermal layout.</p></body></html>")
    server.start()
    try:
        archive.ingest_url(server.url, title="Thermal layout website", source_type="website", collection="web")
    finally:
        server.stop()

    media_doc = archive.ingest_url(
        "https://www.youtube.com/watch?v=abc123",
        title="Soldering lecture transcript",
        source_type="media",
        transcript="The video transcript explains solder bridge inspection and flux cleanup after rework.",
    )

    assert archive.search("BOOT_MODE reset timing", limit=3)[0]["sourceType"] == "datasheet"
    assert archive.search("reciprocal rank fusion portable manifest", limit=3)[0]["sourceType"] == "book"
    assert archive.search("regulator thermal copper pours", limit=3)[0]["sourceType"] == "website"
    assert archive.search("solder bridge flux cleanup", limit=3)[0]["documentId"] == media_doc.document_id
    dense_only_query = dense_collision_query(["reciprocal", "rank", "fusion", "portable", "manifest"])
    assert archive.search(dense_only_query, limit=3, source_types=["book"], mode="lexical") == []
    dense_fallback = archive.search(dense_only_query, limit=1, source_types=["book"])
    assert dense_fallback[0]["sourceType"] == "book"
    assert dense_fallback[0]["denseScore"] >= DENSE_ONLY_MIN_SCORE

    datasheet_id = archive.search("BOOT_MODE reset timing", limit=1)[0]["documentId"]
    archive.invalidate_document(datasheet_id, state="stale", reason="Superseded by newer vendor revision.")
    assert all(result["documentId"] != datasheet_id for result in archive.search("BOOT_MODE reset timing", limit=10))
    stale_results = archive.search("BOOT_MODE reset timing", limit=10, states=["stale"])
    assert stale_results and stale_results[0]["documentId"] == datasheet_id

    archive.remove_document(media_doc.document_id)
    assert archive.search("solder bridge flux cleanup", limit=10) == []

    manifest = archive.portable_manifest()
    manifest_paths = {entry["path"] for entry in manifest["files"]}
    assert "catalog.sqlite" in manifest_paths
    assert "indexes/local-hash-64/chunks.tvim" in manifest_paths
    assert any(path.startswith("snapshots/") for path in manifest_paths)
    assert manifest["turbovecVersion"] == version("turbovec")
    assert manifest["turbovecIndexFormat"] == "tvim"
    assert manifest["denseOnlyMinScore"] == DENSE_ONLY_MIN_SCORE
    index_manifest = json.loads((tmp_path / "archive" / "indexes" / "local-hash-64" / "manifest.json").read_text(encoding="utf-8"))
    assert index_manifest["turbovecVersion"] == version("turbovec")
    assert index_manifest["turbovecIndexFormat"] == "tvim"
    assert index_manifest["denseOnlyMinScore"] == DENSE_ONLY_MIN_SCORE

    restored = DocumentationArchive(tmp_path / "archive")
    assert restored.search("reciprocal rank fusion", limit=1)[0]["sourceType"] == "book"


def test_archive_stats_reports_storage_totals(tmp_path: Path) -> None:
    archive = DocumentationArchive(tmp_path / "archive")
    snapshot_dir = tmp_path / "archive" / "snapshots" / "manual"
    artifact_dir = snapshot_dir / "extracted" / "tables"
    artifact_dir.mkdir(parents=True)
    (snapshot_dir / "source.md").write_text("Snapshot bytes for size accounting.\n", encoding="utf-8")
    (artifact_dir / "table.csv").write_text("name,value\nalpha,1\n", encoding="utf-8")
    archive.rebuild_index()

    stats = archive.stats()
    manifest = archive.portable_manifest()
    files = stats["portableFiles"]
    by_path = {entry["path"]: entry for entry in files}
    archive_size = stats["archiveSize"]

    assert manifest["archiveSize"] == archive_size
    assert archive_size["fileCount"] == len(files)
    assert archive_size["logicalBytes"] == sum(entry["bytes"] for entry in files)
    if archive_size["allocatedBytesAvailable"]:
        assert archive_size["allocatedBytes"] == sum(entry["allocatedBytes"] for entry in files)
    else:
        assert archive_size["allocatedBytes"] is None
    assert by_path["catalog.sqlite"]["category"] == "database"
    assert by_path["snapshots/manual/source.md"]["category"] == "snapshot"
    assert by_path["snapshots/manual/extracted/tables/table.csv"]["category"] == "artifact"
    assert by_path["indexes/local-hash-64/chunks.tvim"]["category"] == "index"
    assert archive_size["databaseBytes"] == (tmp_path / "archive" / "catalog.sqlite").stat().st_size
    assert archive_size["snapshotBytes"] >= by_path["snapshots/manual/source.md"]["bytes"]
    assert archive_size["artifactBytes"] == by_path["snapshots/manual/extracted/tables/table.csv"]["bytes"]
    assert archive_size["indexBytes"] >= by_path["indexes/local-hash-64/chunks.tvim"]["bytes"]
    assert archive_size["denseIndexBytes"] == by_path["indexes/local-hash-64/chunks.tvim"]["bytes"]
    assert archive_size["runtimeEstimateBytes"] == archive_size["denseIndexBytes"]
    assert archive_size["runtimeEstimateKind"] == "dense-index-file"


def test_archive_locality_survives_archive_root_move(tmp_path: Path) -> None:
    archive_root = tmp_path / "archive"
    archive = DocumentationArchive(archive_root)
    document = archive.ingest_text(
        title="Migrated archive note",
        text="Moved archive still searches for MIGRATE-LOCALITY-7.",
        uri="manual://migrated-locality",
    )

    migrated_root = tmp_path / "migrated-archive"
    shutil.copytree(archive_root, migrated_root)
    migrated = DocumentationArchive(migrated_root)
    locality = migrated.stats()["archiveLocality"]

    assert locality["ok"] is True
    assert locality["checkedPathCount"] >= 4
    assert locality["violations"] == []
    assert migrated.search("MIGRATE-LOCALITY-7", limit=1)[0]["documentId"] == document.document_id


def test_archive_locality_reports_absolute_and_outside_snapshot_paths(tmp_path: Path) -> None:
    archive = DocumentationArchive(tmp_path / "archive")
    absolute_document = archive.ingest_text(
        title="Absolute snapshot path",
        text="Absolute snapshot paths should be reported.",
        uri="manual://absolute-snapshot",
    )
    outside_document = archive.ingest_text(
        title="Outside snapshot path",
        text="Outside relative snapshot paths should be reported.",
        uri="manual://outside-snapshot",
    )
    with archive._connect() as db:
        db.execute("UPDATE documents SET snapshot_path = ? WHERE document_id = ?", (str(tmp_path / "outside.md"), absolute_document.document_id))
        db.execute("UPDATE documents SET snapshot_path = ? WHERE document_id = ?", ("../outside.md", outside_document.document_id))

    locality = archive.locality_report()

    assert locality["ok"] is False
    assert {
        (violation["documentId"], violation["reason"])
        for violation in locality["violations"]
        if violation["kind"] == "document-snapshot"
    } == {
        (absolute_document.document_id, "stored path must be relative to archiveRoot"),
        (outside_document.document_id, "stored path resolves outside archiveRoot"),
    }


def test_archive_locality_reports_unsafe_artifact_paths(tmp_path: Path) -> None:
    archive = DocumentationArchive(tmp_path / "archive")
    document = archive.ingest_text(
        title="Unsafe artifact path",
        text="Artifact metadata should stay below the extracted artifact directory.",
        uri="manual://unsafe-artifact",
    )
    stored_snapshot_path = archive.get_document(document.document_id)["snapshot_path"]
    snapshot_path = tmp_path / "archive" / stored_snapshot_path
    artifact_root = snapshot_path.parent / "extracted"
    artifact_root.mkdir()
    (artifact_root / "image_metadata.json").write_text(json.dumps({"artifacts": ["../outside.png"]}), encoding="utf-8")

    locality = archive.locality_report()

    assert locality["ok"] is False
    assert locality["violations"] == [
        {
            "kind": "document-artifact",
            "path": stored_snapshot_path,
            "reason": "Document artifact path must stay inside the extracted artifact directory.",
            "documentId": document.document_id,
        }
    ]


def test_archive_lists_documents_with_pagination_filters_and_order(tmp_path: Path) -> None:
    archive = DocumentationArchive(tmp_path / "archive")
    alpha = archive.ingest_text(title="Alpha Manual", text="Alpha board calibration notes.", uri="manual://alpha", collection="board")
    beta = archive.ingest_text(title="Beta Manual", text="Beta board reset notes.", uri="manual://beta", collection="board")
    gamma = archive.ingest_text(title="Gamma Note", text="Gamma operations note.", uri="manual://gamma", collection="ops")
    with archive._connect() as db:
        db.execute("UPDATE documents SET updated_at = ? WHERE document_id = ?", ("2026-01-01T00:00:00Z", alpha.document_id))
        db.execute("UPDATE documents SET updated_at = ? WHERE document_id = ?", ("2026-01-02T00:00:00Z", beta.document_id))
        db.execute("UPDATE documents SET updated_at = ? WHERE document_id = ?", ("2026-01-03T00:00:00Z", gamma.document_id))

    first_page = archive.list_document_page(limit=2, offset=0, sort_direction="desc")
    assert [document["title"] for document in first_page["documents"]] == ["Gamma Note", "Beta Manual"]
    assert first_page["window"] == {"offset": 0, "limit": 2, "total": 3, "hasMore": True}

    second_page = archive.list_document_page(limit=2, offset=2, sort_direction="desc")
    assert [document["title"] for document in second_page["documents"]] == ["Alpha Manual"]
    assert second_page["window"] == {"offset": 2, "limit": 2, "total": 3, "hasMore": False}

    filtered = archive.list_document_page(limit=10, query="manual://beta", collection="board", sort_direction="asc")
    assert [document["document_id"] for document in filtered["documents"]] == [beta.document_id]
    assert filtered["window"] == {"offset": 0, "limit": 10, "total": 1, "hasMore": False}

    with pytest.raises(ArchiveError, match="limit must be between"):
        archive.list_document_page(limit=0)
    with pytest.raises(ArchiveError, match="sort_direction must be asc or desc"):
        archive.list_document_page(sort_direction="sideways")


def test_fastapi_surface_controls_archive(tmp_path: Path) -> None:
    app = create_app(tmp_path / "archive")
    client = TestClient(app)

    health = client.get("/health")
    assert health.status_code == 200
    assert health.json()["portable"] is True
    assert health.json()["archiveLocality"]["ok"] is True

    ingested = client.post(
        "/ingest/text",
        json={
            "title": "Manual text source",
            "text": "The manual documents a calibration register and ADC offset workflow.",
            "uri": "manual://adc",
            "sourceType": "datasheet",
        },
    )
    assert ingested.status_code == 200
    document_id = ingested.json()["document"]["documentId"]

    search = client.post("/search", json={"query": "ADC calibration register", "limit": 5})
    assert search.status_code == 200
    assert search.json()["results"][0]["documentId"] == document_id

    listed = client.get("/documents", params={"limit": 1, "offset": 0, "query": "manual", "sortDirection": "asc"})
    assert listed.status_code == 200
    assert listed.json()["documents"][0]["document_id"] == document_id
    assert listed.json()["window"] == {"offset": 0, "limit": 1, "total": 1, "hasMore": False}

    stats = client.get("/stats")
    manifest = client.get("/portable-manifest")
    assert stats.status_code == 200
    assert manifest.status_code == 200
    assert stats.json()["archiveLocality"]["ok"] is True
    assert stats.json()["archiveSize"] == manifest.json()["archiveSize"]
    assert stats.json()["archiveSize"]["fileCount"] == len(manifest.json()["files"])
    assert stats.json()["archiveSize"]["logicalBytes"] == sum(entry["bytes"] for entry in manifest.json()["files"])

    invalidated = client.post(
        "/invalidate",
        json={"documentId": document_id, "state": "revoked", "reason": "Incorrect register map."},
    )
    assert invalidated.status_code == 200
    assert invalidated.json()["document"]["state"] == "revoked"
    assert client.post("/search", json={"query": "ADC calibration register"}).json()["results"] == []


def test_ai_enrichment_adds_searchable_provenance_and_replaces_prior_ai_chunks(tmp_path: Path) -> None:
    archive = DocumentationArchive(tmp_path / "archive")
    document = archive.ingest_text(
        title="Visual import note",
        text="The PDF extraction captured a POWER table and a reset sequencing flowchart.",
        source_type="datasheet",
        uri="mock://visual-import",
    )

    enriched = archive.enrich_document(
        document.document_id,
        spans=[
            archive_module.ExtractedSpan("AI visual summary says FLOWCHART-ENRICH-1 has RESET LOW then ENABLE RAIL.", "ai:visual:flowchart"),
            archive_module.ExtractedSpan("AI metadata summary says the source has a power sequencing table.", "ai:metadata:tables"),
        ],
        model="gpt-test",
        skill_ids=["documentation-enrich-visuals"],
        summary="Added visual extraction notes.",
        payload={"warning": "none"},
    )

    ai_chunks = [chunk for chunk in enriched["chunks"] if chunk["chunk_origin"] == "ai"]
    assert len(ai_chunks) == 2
    assert {chunk["locator"] for chunk in ai_chunks} == {"ai:visual:flowchart", "ai:metadata:tables"}
    assert enriched["enrichments"][0]["model"] == "gpt-test"

    result = archive.search("FLOWCHART-ENRICH-1 RESET LOW", limit=1)[0]
    assert result["documentId"] == document.document_id
    assert result["chunkOrigin"] == "ai"
    assert result["enrichmentId"] == ai_chunks[0]["enrichment_id"]

    archive.enrich_document(
        document.document_id,
        spans=[archive_module.ExtractedSpan("AI rerun summary says FLOWCHART-ENRICH-2 replaced the prior visual note.", "ai:visual:rerun")],
        model="gpt-test",
        skill_ids=["documentation-enrich-visuals"],
        summary="Replaced visual extraction notes.",
        payload={},
    )

    record = archive.get_document(document.document_id)
    assert [chunk["locator"] for chunk in record["chunks"] if chunk["chunk_origin"] == "ai"] == ["ai:visual:rerun"]
    assert archive.search("FLOWCHART-ENRICH-1", limit=10) == []
    assert archive.search("FLOWCHART-ENRICH-2", limit=1)[0]["chunkOrigin"] == "ai"


def test_fastapi_enrich_endpoint_writes_derived_chunks(tmp_path: Path) -> None:
    app = create_app(tmp_path / "archive")
    client = TestClient(app)
    ingested = client.post(
        "/ingest/text",
        json={"title": "Media note", "text": "Transcript source mentions a setup screen.", "sourceType": "media"},
    )
    document_id = ingested.json()["document"]["documentId"]

    response = client.post(
        f"/documents/{document_id}/enrich",
        json={
            "model": "gpt-test",
            "skillIds": ["documentation-enrich-media"],
            "summary": "Added media sections.",
            "spans": [{"locator": "ai:media:0", "text": "AI media summary says MEDIA-ENRICH-7 appears in the setup section."}],
            "payload": {"source": "test"},
        },
    )

    assert response.status_code == 200
    assert response.json()["document"]["enrichments"][0]["skill_ids_json"] == '["documentation-enrich-media"]'
    search = client.post("/search", json={"query": "MEDIA-ENRICH-7 setup section"})
    assert search.json()["results"][0]["chunkOrigin"] == "ai"


def test_ingest_autodetects_title_collection_uri_and_source_type(tmp_path: Path) -> None:
    archive = DocumentationArchive(tmp_path / "archive")

    text_document = archive.ingest_text(text="Autodetected Manual\nThe AXI-77 register controls burst length.")
    text_record = archive.get_document(text_document.document_id)
    assert text_record["title"] == "Autodetected Manual"
    assert text_record["source_type"] == "text"
    assert text_record["uri"].startswith("manual://Autodetected_Manual-")
    assert text_record["collection"] == "manual"
    assert archive.search("AXI-77 burst length", collection="manual", limit=1)[0]["documentId"] == text_document.document_id

    upload_document = archive.ingest_upload(
        filename="uploaded-note.md",
        content=b"Uploaded metadata note says AUTO-UPLOAD-9.",
        content_type="text/markdown",
    )
    upload_record = archive.get_document(upload_document.document_id)
    assert upload_record["title"] == "uploaded-note.md"
    assert upload_record["collection"] == "uploads"

    docs_dir = tmp_path / "vendor-docs"
    docs_dir.mkdir()
    note = docs_dir / "board-guide.md"
    note.write_text("Board guide says AUTO-PATH-12 configures strap pins.\n", encoding="utf-8")
    path_document = archive.ingest_path(note)[0]
    path_record = archive.get_document(path_document.document_id)
    assert path_record["title"] == "board-guide.md"
    assert path_record["collection"] == "vendor-docs"

    server = HtmlFixtureServer("<html><body><h1>Thermal Layout Note</h1><p>AUTO-URL-33 improves thermal layout.</p></body></html>")
    server.start()
    try:
        url_document = archive.ingest_url(server.url)
    finally:
        server.stop()
    url_record = archive.get_document(url_document.document_id)
    assert url_record["title"] == "thermal.html"
    assert url_record["collection"].startswith("127.0.0.1:")


def test_youtube_playlist_url_ingests_each_video_transcript(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    playlist_url = "https://www.youtube.com/playlist?list=PLbringup"
    playlist = archive_module.YouTubePlaylist(
        title="Board Bringup Playlist",
        entries=[
            archive_module.YouTubePlaylistEntry("Reset Sequencing", "video-one", "https://www.youtube.com/watch?v=video-one"),
            archive_module.YouTubePlaylistEntry("Power Rail Checks", "video-two", "https://www.youtube.com/watch?v=video-two"),
        ],
    )
    transcripts = {
        "https://www.youtube.com/watch?v=video-one": "Reset transcript explains PLAYLIST-RESET-1.",
        "https://www.youtube.com/watch?v=video-two": "Power transcript explains PLAYLIST-POWER-2.",
    }
    monkeypatch.setattr(archive_module, "extract_youtube_playlist", lambda url: playlist)
    stub_youtube_media(monkeypatch, transcripts)

    archive = DocumentationArchive(tmp_path / "archive")
    documents = archive.ingest_url_documents(playlist_url)

    assert [document.title for document in documents] == ["Reset Sequencing", "Power Rail Checks"]
    records = [archive.get_document(document.document_id) for document in documents]
    assert {record["collection"] for record in records} == {"Board Bringup Playlist"}
    assert {record["source_type"] for record in records} == {"media"}
    for record in records:
        assert "description" in {chunk["locator"] for chunk in record["chunks"]}
        snapshot = tmp_path / "archive" / record["snapshot_path"]
        assert (snapshot.parent / "extracted" / "media" / "keyframes.tsv").exists()
        assert (snapshot.parent / "extracted" / "media" / "keyframes" / "frame-000001.jpg").exists()
    assert archive.search("PLAYLIST-POWER-2", collection="Board Bringup Playlist", limit=1)[0]["title"] == "Power Rail Checks"
    assert archive.search("Metadata for video-one allocator pressure slides", collection="Board Bringup Playlist", limit=1)[0]["title"] == "Reset Sequencing"
    assert archive.search("Metadata for video-two allocator pressure slides", collection="Board Bringup Playlist", limit=1)[0]["title"] == "Power Rail Checks"
    assert archive.search("selected YouTube slide frame", collection="Board Bringup Playlist", limit=2)

    app = create_app(tmp_path / "api-archive")
    client = TestClient(app)
    response = client.post("/ingest/url", json={"url": playlist_url})

    assert response.status_code == 200
    payload = response.json()
    assert payload["document"]["title"] == "Reset Sequencing"
    assert [document["title"] for document in payload["documents"]] == ["Reset Sequencing", "Power Rail Checks"]


def test_youtube_url_streams_progress_events(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    video_url = "https://www.youtube.com/watch?v=stream-demo"
    stub_youtube_media(monkeypatch, {video_url: "Transcript explains STREAM-PROGRESS-1."})
    app = create_app(tmp_path / "api-archive")
    client = TestClient(app)

    with client.stream("POST", "/ingest/url?stream=1", json={"url": video_url}) as response:
        lines = [json.loads(line) for line in response.iter_lines() if line]

    assert response.status_code == 200
    assert any(line.get("type") == "progress" and "transcript" in line.get("stage", "").lower() for line in lines)
    assert lines[-1]["type"] == "result"
    assert lines[-1]["result"]["document"]["title"] == "Video stream-demo"


def test_youtube_evidence_extracts_transcript_and_keyframes_in_parallel(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    metadata = archive_module.YouTubeVideoMetadata(
        title="Parallel video",
        webpage_url="https://www.youtube.com/watch?v=parallel",
        stream_url="https://example.com/parallel.mp4",
        http_headers={},
        duration=120,
    )
    events: list[str] = []
    keyframe_started = threading.Event()

    def transcribe(_url: str, _metadata: archive_module.YouTubeVideoMetadata, *, progress=None) -> list[archive_module.TranscriptSegment]:
        events.append("transcript-start")
        assert keyframe_started.wait(1)
        events.append("transcript-finish")
        return [archive_module.TranscriptSegment(0.0, 30.0, "parallel transcript")]

    def keyframes(_metadata: archive_module.YouTubeVideoMetadata, artifact_dir: Path, *, transcript_segments=None, progress=None) -> list[dict[str, object]]:
        events.append("keyframe-start")
        (artifact_dir / "media").mkdir(parents=True, exist_ok=True)
        keyframe_started.set()
        time.sleep(0.02)
        events.append("keyframe-finish")
        return [{"offsetSeconds": 0, "path": "media/keyframes/frame-000001.jpg", "reason": "segment-start"}]

    monkeypatch.setattr(archive_module, "transcribe_youtube_video", transcribe)
    monkeypatch.setattr(archive_module, "capture_youtube_keyframes", keyframes)

    segments, selected = archive_module.extract_youtube_video_evidence("https://www.youtube.com/watch?v=parallel", metadata, tmp_path / "artifact")

    assert [segment.text for segment in segments] == ["parallel transcript"]
    assert events.index("keyframe-start") < events.index("transcript-finish")
    assert selected[0]["transcriptStartSeconds"] == 0.0
    assert selected[0]["transcriptEndSeconds"] == 30.0
    assert (tmp_path / "artifact" / "media" / "keyframes.tsv").exists()


def test_youtube_video_ingest_preserves_transcript_metadata_and_keyframes(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    video_url = "https://www.youtube.com/watch?v=visual-demo"
    stub_youtube_media(
        monkeypatch,
        {video_url: "Transcript explains YOUTUBE-VISUAL-9 while the slides show allocator pressure."},
    )
    archive = DocumentationArchive(tmp_path / "archive")

    document = archive.ingest_url(video_url)
    record = archive.get_document(document.document_id)
    snapshot = tmp_path / "archive" / record["snapshot_path"]
    extracted = snapshot.parent / "extracted"

    assert record["title"] == "Video visual-demo"
    assert record["source_type"] == "media"
    description_chunk = next(chunk for chunk in record["chunks"] if chunk["locator"] == "description")
    metadata_chunk = next(chunk for chunk in record["chunks"] if chunk["locator"] == "media metadata")
    assert "YouTube video description:" in description_chunk["text"]
    assert "Metadata for visual-demo includes allocator pressure slides." in description_chunk["text"]
    assert "Metadata for visual-demo includes allocator pressure slides." not in metadata_chunk["text"]
    assert archive.search("YOUTUBE-VISUAL-9 allocator pressure", limit=1)[0]["documentId"] == document.document_id
    assert archive.search("Duration seconds 1234", limit=1)[0]["documentId"] == document.document_id
    assert archive.search("Metadata for visual-demo allocator pressure slides", limit=1)[0]["documentId"] == document.document_id
    assert (extracted / "media" / "youtube_metadata.json").exists()
    assert json.loads((extracted / "media" / "youtube_metadata.json").read_text(encoding="utf-8"))["description"] == "Metadata for visual-demo includes allocator pressure slides."
    assert (extracted / "media" / "description.txt").read_text(encoding="utf-8") == "Metadata for visual-demo includes allocator pressure slides.\n"
    assert (extracted / "media" / "keyframes.tsv").read_text(encoding="utf-8").splitlines() == [
        "offset_seconds\tpath\treason\tchange_score\ttranscript_start_seconds\ttranscript_end_seconds",
        "0\tmedia/keyframes/frame-000001.jpg\tsegment-start\t\t0.0\t2.0",
        "2\tmedia/keyframes/frame-000002.jpg\tvisual-change\t0.5\t0.0\t2.0",
    ]
    keyframe_chunks = [chunk for chunk in record["chunks"] if chunk["locator"].startswith("media keyframe ")]
    assert [chunk["locator"] for chunk in keyframe_chunks] == [
        "media keyframe keyframe-000000 00:00",
        "media keyframe keyframe-000002 00:02",
    ]
    assert "Artifact path: media/keyframes/frame-000001.jpg." in keyframe_chunks[0]["text"]
    assert "Selected YouTube slide frame keyframe-000000 at 00:00 (0 seconds)." in keyframe_chunks[0]["text"]
    assert "Transcript near this frame:" in keyframe_chunks[0]["text"]
    assert "[00:00 -> 00:02] Transcript explains YOUTUBE-VISUAL-9 while the slides show allocator pressure." in keyframe_chunks[0]["text"]
    assert "Artifact path: media/keyframes/frame-000002.jpg." in keyframe_chunks[1]["text"]
    assert "Selection reason: visual-change." in keyframe_chunks[1]["text"]
    assert "media keyframes" not in {chunk["locator"] for chunk in record["chunks"]}
    artifacts = record["artifacts"]
    assert [artifact["locator"] for artifact in artifacts] == [
        "media keyframe keyframe-000000 00:00",
        "media keyframe keyframe-000002 00:02",
    ]
    assert (extracted / "media" / "transcript_segments.tsv").exists()
    assert (extracted / "media" / "visual_sampling.json").exists()
    assert (extracted / "media" / "keyframes" / "frame-000001.jpg").exists()
    assert (extracted / "media" / "keyframes" / "frame-000002.jpg").exists()


def test_document_detail_supports_chunk_windows_and_truncation(tmp_path: Path) -> None:
    app = create_app(tmp_path / "windowed-archive")
    client = TestClient(app)
    text = "\n\n".join(f"Section {index} " + "alpha " * 260 for index in range(4))
    ingest_response = client.post("/ingest/text", json={"title": "Large source", "text": text})
    document_id = ingest_response.json()["document"]["documentId"]

    response = client.get(
        f"/documents/{document_id}",
        params={
            "chunkOffset": 1,
            "chunkLimit": 1,
            "chunkTextMaxChars": 40,
            "artifactOffset": 0,
            "artifactLimit": 0,
        },
    )

    assert response.status_code == 200
    document = response.json()["document"]
    assert len(document["chunks"]) == 1
    assert document["chunks"][0]["textTruncated"] is True
    assert document["chunks"][0]["textLength"] > 40
    assert len(document["chunks"][0]["text"]) <= 43
    assert document["chunkWindow"]["offset"] == 1
    assert document["chunkWindow"]["limit"] == 1
    assert document["chunkWindow"]["total"] > 1
    assert document["chunkWindow"]["hasMore"] is True
    assert document["artifacts"] == []
    assert document["artifactWindow"] == {"offset": 0, "limit": 0, "total": 0, "hasMore": False}


def test_youtube_keyframe_spans_include_timestamped_transcript_context() -> None:
    keyframes = [
        {
            "offsetSeconds": 12,
            "path": "media/keyframes/frame-000001.png",
            "reason": "visual-change",
            "changeScore": 0.42,
            "transcriptStartSeconds": 10.0,
            "transcriptEndSeconds": 20.0,
        }
    ]
    transcript_segments = [
        archive_module.TranscriptSegment(8.0, 9.5, "before the selected slide"),
        archive_module.TranscriptSegment(10.0, 14.0, "the slide explains KEYFRAME-CONTEXT-12"),
        archive_module.TranscriptSegment(21.0, 22.0, "after the selected slide"),
    ]

    spans = archive_module.youtube_keyframe_spans(keyframes, transcript_segments)

    assert len(spans) == 1
    assert spans[0].locator == "media keyframe keyframe-000012 00:12"
    assert "Selected YouTube slide frame keyframe-000012 at 00:12 (12 seconds)." in spans[0].text
    assert "Artifact path: media/keyframes/frame-000001.png." in spans[0].text
    assert "Transcript window: 00:10-00:20." in spans[0].text
    assert "KEYFRAME-CONTEXT-12" in spans[0].text
    assert "before the selected slide" not in spans[0].text


def test_slide_frame_selector_keeps_one_frame_per_visual_state(tmp_path: Path) -> None:
    frame_paths = []
    for index, color in enumerate(["white", "white", "white", "black", "black", "black"], start=1):
        path = tmp_path / f"frame-{index:06d}.jpg"
        Image.new("RGB", (64, 36), color).save(path)
        frame_paths.append(path)

    selected = archive_module.select_slide_frames(
        frame_paths,
        0,
        archive_module.VideoVisualProfile(settle_seconds=1),
    )

    assert [(candidate["offsetSeconds"], candidate["reason"]) for candidate in selected] == [
        (0, "segment-start"),
        (4, "visual-change"),
    ]


def test_youtube_keyframe_capture_scans_downloaded_local_video(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    metadata = archive_module.YouTubeVideoMetadata(
        title="Long presentation",
        webpage_url="https://www.youtube.com/watch?v=local-scan",
        stream_url="https://video.example/remote.mp4",
        http_headers={"User-Agent": "cloudx-test"},
        duration=601,
    )
    visual_path = tmp_path / "downloaded-video.mp4"
    download_calls: list[str] = []
    scan_calls: list[tuple[str, object, int, int | None]] = []

    def download_once(
        metadata_arg: archive_module.YouTubeVideoMetadata,
        _output_dir: Path,
        *,
        progress,
        started_at: float,
        profile: archive_module.VideoVisualProfile,
    ) -> Path:
        download_calls.append(metadata_arg.webpage_url)
        visual_path.write_bytes(b"downloaded video")
        return visual_path

    def scan_local(input_url: str, http_headers, start: int, duration: int | None, output_dir: Path, profile: archive_module.VideoVisualProfile) -> dict[str, object]:
        scan_calls.append((input_url, http_headers, start, duration))
        output_dir.mkdir(parents=True, exist_ok=True)
        colors = {0: "white", 300: "black", 600: "blue"}
        frame_path = output_dir / "frame-000001.jpg"
        Image.new("RGB", (64, 36), colors[start]).save(frame_path)
        return {
            "scannedFrames": 1,
            "selected": [{
                "offsetSeconds": start,
                "sourcePath": frame_path,
                "reason": "segment-start",
                "changeScore": 0.0,
            }],
        }

    monkeypatch.setattr(archive_module, "download_youtube_visual_source", download_once)
    monkeypatch.setattr(archive_module, "scan_video_segment", scan_local)

    keyframes = archive_module.capture_youtube_keyframes(
        metadata,
        tmp_path / "artifact",
        profile=archive_module.VideoVisualProfile(segment_seconds=300, local_workers=2),
    )

    assert download_calls == ["https://www.youtube.com/watch?v=local-scan"]
    assert sorted(scan_calls, key=lambda call: call[2]) == [
        (str(visual_path), None, 0, 300),
        (str(visual_path), None, 300, 300),
        (str(visual_path), None, 600, 1),
    ]
    assert [keyframe["offsetSeconds"] for keyframe in keyframes] == [0, 300, 600]
    manifest = json.loads((tmp_path / "artifact" / "media" / "visual_sampling.json").read_text(encoding="utf-8"))
    assert manifest["strategy"] == "downloaded-slide-change"
    assert manifest["workers"] == 2
    assert manifest["selectedFrames"] == 3


def test_youtube_audio_download_progress_reports_bytes_and_eta() -> None:
    events: list[dict[str, object]] = []
    hook = archive_module.youtube_audio_download_progress_hook(events.append, time.monotonic(), 29450)

    hook({"status": "downloading", "downloaded_bytes": 50, "total_bytes": 200, "eta": 12})
    hook({"status": "finished", "downloaded_bytes": 200})

    assert events[0]["stage"] == "Downloading YouTube audio for local transcription."
    assert events[0]["progress"] == 27
    assert events[0]["etaSeconds"] == 12
    assert events[0]["metrics"] == {"durationSeconds": 29450, "downloadedBytes": 50, "totalBytes": 200}
    assert events[-1]["stage"] == "Finished downloading YouTube audio."
    assert events[-1]["progress"] == 31
    assert events[-1]["metrics"] == {"durationSeconds": 29450, "downloadedBytes": 200}


def test_youtube_audio_download_progress_omits_unstable_early_eta() -> None:
    events: list[dict[str, object]] = []
    hook = archive_module.youtube_audio_download_progress_hook(events.append, time.monotonic(), 29450)

    hook({"status": "downloading", "downloaded_bytes": 1024, "total_bytes": 386095778})

    assert "etaSeconds" not in events[0]
    assert events[0]["metrics"] == {"durationSeconds": 29450, "downloadedBytes": 1024, "totalBytes": 386095778}


def test_documentation_asr_backend_is_explicit(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("CLOUDX_DOCUMENTATION_ASR_BACKEND", "whisper_cpp")
    assert archive_module.documentation_asr_backend() == "whisper-cpp"

    monkeypatch.setenv("CLOUDX_DOCUMENTATION_ASR_BACKEND", "openvino")
    with pytest.raises(ArchiveError, match="Unsupported documentation ASR backend"):
        archive_module.documentation_asr_backend()


def test_whisper_cpp_extra_args_are_explicit(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.delenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_ARGS", raising=False)

    assert archive_module.documentation_whisper_cpp_extra_args() == []
    assert archive_module.documentation_whisper_cpp_stability_args() == ["-sns", "-nf", "-mc", "0"]

    monkeypatch.setenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_ARGS", "--print-colors")
    assert archive_module.documentation_whisper_cpp_extra_args() == ["--print-colors"]


def test_whisper_cpp_vad_requires_configured_model(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_VAD", "true")
    monkeypatch.delenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_VAD_MODEL_PATH", raising=False)
    monkeypatch.delenv("CLOUDX_ASR_WHISPER_CPP_VAD_MODEL_PATH", raising=False)

    with pytest.raises(ArchiveError, match="VAD_MODEL_PATH is required"):
        archive_module.documentation_whisper_cpp_vad_args()


def test_whisper_cpp_backend_converts_invokes_cli_and_parses_segments(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    audio_path = tmp_path / "audio.webm"
    audio_path.write_bytes(b"audio")
    model_path = tmp_path / "ggml-large-v3.bin"
    model_path.write_bytes(b"model")
    vad_model_path = tmp_path / "ggml-silero-v6.2.0.bin"
    vad_model_path.write_bytes(b"vad")
    binary_path = tmp_path / "whisper-cli"
    binary_path.write_text("#!/bin/sh\n", encoding="utf-8")
    binary_path.chmod(0o755)
    events: list[dict[str, object]] = []
    commands: list[list[str]] = []

    class FakeProcess:
        def __init__(self, command, **_kwargs):
            commands.append(command)
            if command[0] == "ffmpeg":
                Path(command[-1]).write_bytes(b"wav")
                self.stdout = iter(["out_time_ms=30000000\n", "progress=continue\n", "out_time_ms=60000000\n", "progress=end\n"])
            else:
                output_base = Path(command[command.index("-of") + 1])
                output_base.with_suffix(".json").write_text(
                    json.dumps(
                        {
                            "transcription": [
                                {
                                    "offsets": {"from": 1230, "to": 4560},
                                    "text": " GPU accelerated transcript ",
                                }
                            ]
                        }
                    ),
                    encoding="utf-8",
                )
                self.stdout = iter(["whisper_print_progress_callback: progress = 25%\n", "progress = 100%\n"])

        def wait(self):
            return 0

    monkeypatch.setattr(archive_module.subprocess, "Popen", FakeProcess)
    monkeypatch.setenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_MODEL_PATH", str(model_path))
    monkeypatch.setenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_BIN", str(binary_path))
    monkeypatch.setenv("CLOUDX_DOCUMENTATION_ASR_BEAM_SIZE", "1")
    monkeypatch.setenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_THREADS", "3")
    monkeypatch.setenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_VAD", "true")
    monkeypatch.setenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_VAD_MODEL_PATH", str(vad_model_path))
    metadata = archive_module.YouTubeVideoMetadata(
        title="GPU Demo",
        webpage_url="https://www.youtube.com/watch?v=gpu",
        stream_url="https://example.com/gpu.mp4",
        http_headers={},
        duration=60,
    )

    segments = archive_module.transcribe_audio_whisper_cpp(audio_path, tmp_path, metadata, audio_path.stat().st_size, time.monotonic(), progress=events.append)

    assert segments == [archive_module.TranscriptSegment(1.23, 4.56, "GPU accelerated transcript")]
    assert commands[0][:5] == ["ffmpeg", "-hide_banner", "-loglevel", "error", "-nostdin"]
    assert commands[0][commands[0].index("-progress") + 1] == "pipe:1"
    assert commands[0][commands[0].index("-stats_period") + 1] == "5"
    assert commands[1][0] == str(binary_path)
    assert commands[1][commands[1].index("-m") + 1] == str(model_path)
    assert commands[1][commands[1].index("-f") + 1] == str(tmp_path / "whisper-cpp-input.wav")
    assert "-oj" in commands[1]
    assert commands[1][commands[1].index("-bs") + 1] == "1"
    assert commands[1][commands[1].index("-t") + 1] == "3"
    assert commands[1][commands[1].index("-sns") + 1] == "-nf"
    assert commands[1][commands[1].index("-mc") + 1] == "0"
    assert commands[1][commands[1].index("--vad-model") + 1] == str(vad_model_path)
    assert any("Converted YouTube audio through 00:30" in str(event["stage"]) for event in events)
    assert any("25% complete" in str(event["stage"]) for event in events)


def test_whisper_cpp_backend_chunks_long_audio_and_offsets_segments(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    audio_path = tmp_path / "audio.webm"
    audio_path.write_bytes(b"audio")
    model_path = tmp_path / "ggml-large-v3.bin"
    model_path.write_bytes(b"model")
    binary_path = tmp_path / "whisper-cli"
    binary_path.write_text("#!/bin/sh\n", encoding="utf-8")
    binary_path.chmod(0o755)
    commands: list[list[str]] = []
    split_commands: list[list[str]] = []
    events: list[dict[str, object]] = []

    def fake_conversion(command, *_args, **_kwargs):
        Path(command[-1]).write_bytes(b"wav")

    def fake_run(command, **_kwargs):
        split_commands.append(command)
        output_path = Path(command[-1])
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b"chunk")
        return subprocess.CompletedProcess(command, 0, "", "")

    class FakeProcess:
        def __init__(self, command, **_kwargs):
            commands.append(command)
            output_base = Path(command[command.index("-of") + 1])
            chunk_path = Path(command[command.index("-f") + 1])
            chunk_number = int(chunk_path.stem.rsplit("-", 1)[1])
            start_milliseconds = 1000 if chunk_number == 0 else 6000
            output_base.with_suffix(".json").write_text(
                json.dumps(
                    {
                        "transcription": [
                            {
                                "offsets": {"from": start_milliseconds, "to": start_milliseconds + 1000},
                                "text": f" transcript for {chunk_path.stem} ",
                            }
                        ]
                    }
                ),
                encoding="utf-8",
            )
            self.stdout = iter(["progress = 0%\n", "progress = 100%\n"])

        def wait(self):
            return 0

    monkeypatch.setattr(archive_module, "run_ffmpeg_audio_conversion", fake_conversion)
    monkeypatch.setattr(archive_module.subprocess, "run", fake_run)
    monkeypatch.setattr(archive_module.subprocess, "Popen", FakeProcess)
    monkeypatch.setenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_MODEL_PATH", str(model_path))
    monkeypatch.setenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_BIN", str(binary_path))
    monkeypatch.setenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_CHUNK_SECONDS", "60")
    monkeypatch.setenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_CHUNK_OVERLAP_SECONDS", "5")
    monkeypatch.setenv("CLOUDX_DOCUMENTATION_WHISPER_CPP_VAD", "false")
    metadata = archive_module.YouTubeVideoMetadata(
        title="Chunked Demo",
        webpage_url="https://www.youtube.com/watch?v=chunked",
        stream_url="https://example.com/chunked.mp4",
        http_headers={},
        duration=125,
    )

    segments = archive_module.transcribe_audio_whisper_cpp(audio_path, tmp_path, metadata, audio_path.stat().st_size, time.monotonic(), progress=events.append)

    assert [round(segment.start_seconds, 1) for segment in segments] == [1.0, 61.0, 121.0]
    assert [round(segment.end_seconds, 1) for segment in segments] == [2.0, 62.0, 122.0]
    assert len(commands) == 3
    assert len(split_commands) == 3
    assert [(command[command.index("-ss") + 1], command[command.index("-t") + 1]) for command in split_commands] == [
        ("0.000", "65.000"),
        ("55.000", "70.000"),
        ("115.000", "10.000"),
    ]
    assert all(command[command.index("-f") + 1].endswith(f"chunk-{index:04d}.wav") for index, command in enumerate(commands))
    assert any(event.get("metrics", {}).get("chunkOverlapSeconds") == 5 for event in events)
    assert any(event.get("metrics", {}).get("chunksTotal") == 3 for event in events)


def test_whisper_cpp_overlap_merge_keeps_segments_once() -> None:
    left_chunk = archive_module.WhisperCppAudioChunk(
        index=1,
        start_seconds=0.0,
        duration_seconds=65.0,
        keep_start_seconds=0.0,
        keep_end_seconds=60.0,
        path=Path("chunk-0000.wav"),
    )
    right_chunk = archive_module.WhisperCppAudioChunk(
        index=2,
        start_seconds=55.0,
        duration_seconds=70.0,
        keep_start_seconds=60.0,
        keep_end_seconds=120.0,
        path=Path("chunk-0001.wav"),
    )

    left_segments = archive_module.keep_whisper_cpp_chunk_segments(
        [
            archive_module.TranscriptSegment(57.0, 62.0, "sentence crossing the boundary"),
            archive_module.TranscriptSegment(62.0, 64.0, "right overlap duplicate"),
        ],
        left_chunk,
    )
    right_segments = archive_module.keep_whisper_cpp_chunk_segments(
        [
            archive_module.TranscriptSegment(2.0, 7.0, "sentence crossing the boundary"),
            archive_module.TranscriptSegment(6.0, 8.0, "next sentence"),
        ],
        right_chunk,
    )

    assert left_segments == [archive_module.TranscriptSegment(57.0, 62.0, "sentence crossing the boundary")]
    assert right_segments == [archive_module.TranscriptSegment(61.0, 63.0, "next sentence")]


def test_youtube_parallel_progress_reports_channel_percent() -> None:
    events: list[dict[str, object]] = []
    reporter = archive_module.YouTubeParallelProgress(events.append).channel("Transcript", 26, 58)
    assert reporter is not None

    reporter({"stage": "whisper.cpp transcription 50% complete.", "progress": 42})

    assert events == [
        {
            "stage": "Transcript: whisper.cpp transcription 50% complete.",
            "progress": 40,
            "channel": "transcript",
            "channelLabel": "Transcript",
            "channelProgress": 50,
        }
    ]


def test_progress_heartbeat_reports_latest_event() -> None:
    events: list[dict[str, object]] = []
    heartbeat = archive_module.ProgressHeartbeat(
        events.append,
        stage="Running faster-whisper transcription; waiting for timestamped segments.",
        progress=32,
        metrics={"durationSeconds": 29450},
        interval_seconds=0.01,
    )
    try:
        deadline = time.monotonic() + 1
        while not events and time.monotonic() < deadline:
            time.sleep(0.01)
        heartbeat.update(
            stage="Transcribed through 16:27.",
            progress=33,
            eta_seconds=6039,
            metrics={"durationSeconds": 29450, "transcribedSeconds": 986.9},
        )
        deadline = time.monotonic() + 1
        while not any(event.get("stage") == "Transcribed through 16:27." for event in events) and time.monotonic() < deadline:
            time.sleep(0.01)
    finally:
        heartbeat.stop()

    assert events[0] == {
        "stage": "Running faster-whisper transcription; waiting for timestamped segments.",
        "progress": 32,
        "metrics": {"durationSeconds": 29450},
    }
    assert any(
        event == {
            "stage": "Transcribed through 16:27.",
            "progress": 33,
            "etaSeconds": 6039,
            "metrics": {"durationSeconds": 29450, "transcribedSeconds": 986.9},
        }
        for event in events
    )


def test_snapshot_artifact_window_paginates_large_keyframe_index(tmp_path: Path) -> None:
    snapshot = tmp_path / "snapshots" / "video-source.bin"
    media_dir = snapshot.parent / "extracted" / "media"
    frames_dir = media_dir / "keyframes"
    frames_dir.mkdir(parents=True)
    snapshot.parent.mkdir(exist_ok=True)
    snapshot.write_bytes(b"video")
    keyframes = [
        {"offsetSeconds": index, "path": f"media/keyframes/frame-{index + 1:06d}.png"}
        for index in range(25)
    ]
    for index in range(10, 15):
        Image.new("RGB", (64, 36), "white").save(frames_dir / f"frame-{index + 1:06d}.png")
    archive_module.write_keyframe_index(media_dir / "keyframes.tsv", keyframes)

    window = archive_module.snapshot_artifact_window("doc-video", snapshot, offset=10, limit=5)

    assert window.total == 25
    assert [artifact["offsetSeconds"] for artifact in window.artifacts] == [10, 11, 12, 13, 14]
    assert [artifact["locator"] for artifact in window.artifacts] == [
        "media keyframe keyframe-000010 00:10",
        "media keyframe keyframe-000011 00:11",
        "media keyframe keyframe-000012 00:12",
        "media keyframe keyframe-000013 00:13",
        "media keyframe keyframe-000014 00:14",
    ]
    assert all(artifact["available"] is True for artifact in window.artifacts)


def test_failed_youtube_playlist_does_not_leave_active_partial_documents(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    playlist_url = "https://www.youtube.com/playlist?list=PLpartial"
    playlist = archive_module.YouTubePlaylist(
        title="Partial Playlist",
        entries=[
            archive_module.YouTubePlaylistEntry("Good Video", "video-good", "https://www.youtube.com/watch?v=video-good"),
            archive_module.YouTubePlaylistEntry("Bad Video", "video-bad", "https://www.youtube.com/watch?v=video-bad"),
        ],
    )
    monkeypatch.setattr(archive_module, "extract_youtube_playlist", lambda url: playlist)
    stub_youtube_media(monkeypatch, {"https://www.youtube.com/watch?v=video-good": "Good transcript says PLAYLIST-PARTIAL-1."})
    archive = DocumentationArchive(tmp_path / "archive")

    with pytest.raises(KeyError):
        archive.ingest_url_documents(playlist_url)

    assert archive.search("PLAYLIST-PARTIAL-1", limit=10) == []
    deleted = archive.list_documents(states=["deleted"])
    assert [document["title"] for document in deleted] == ["Good Video"]


def test_youtube_video_without_description_omits_empty_description_source(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    video_url = "https://www.youtube.com/watch?v=no-description"
    stub_youtube_media(
        monkeypatch,
        {video_url: "Transcript explains YOUTUBE-NODESC-4 without description text."},
        descriptions={video_url: None},
    )
    archive = DocumentationArchive(tmp_path / "archive")

    document = archive.ingest_url(video_url)
    record = archive.get_document(document.document_id)
    extracted = tmp_path / "archive" / record["snapshot_path"]

    assert "description" not in {chunk["locator"] for chunk in record["chunks"]}
    assert archive.search("YOUTUBE-NODESC-4", limit=1)[0]["documentId"] == document.document_id
    assert not (extracted.parent / "extracted" / "media" / "description.txt").exists()


def test_long_unpunctuated_text_is_chunked_without_dropping_tail(tmp_path: Path) -> None:
    archive = DocumentationArchive(tmp_path / "archive")
    body = " ".join(f"transcriptword{index}" for index in range(900))
    tail = "YTTAIL-900 allocator reclaim pressure watermark"

    document = archive.ingest_text(
        title="Long transcript",
        text=f"{body} {tail}",
        source_type="media",
        uri="mock://long-transcript",
    )

    record = archive.get_document(document.document_id)
    assert len(record["chunks"]) > 1
    assert archive.search("YTTAIL-900 allocator reclaim pressure watermark", limit=1)[0]["documentId"] == document.document_id


def test_reingesting_changed_source_supersedes_old_revision(tmp_path: Path) -> None:
    source = tmp_path / "board-datasheet.md"
    source.write_text("Board Datasheet\nOLDREG reset mode lives at address 0x10.\n", encoding="utf-8")
    archive = DocumentationArchive(tmp_path / "archive")

    old_document = archive.ingest_path(source, source_type="datasheet")
    assert archive.search("OLDREG reset mode", limit=1)[0]["documentId"] == old_document[0].document_id

    source.write_text("Board Datasheet\nNEWREG reset mode lives at address 0x20.\n", encoding="utf-8")
    new_document = archive.ingest_path(source, source_type="datasheet")

    assert new_document[0].document_id != old_document[0].document_id
    assert all(result["documentId"] != old_document[0].document_id for result in archive.search("OLDREG reset mode", limit=10))
    assert archive.search("OLDREG", limit=10, mode="lexical") == []
    new_results = archive.search("NEWREG reset mode", limit=10)
    assert new_results and new_results[0]["documentId"] == new_document[0].document_id
    superseded_results = archive.search("OLDREG reset mode", limit=10, states=["superseded"])
    assert superseded_results and superseded_results[0]["documentId"] == old_document[0].document_id
    old_record = archive.get_document(old_document[0].document_id)
    assert old_record["state"] == "superseded"
    assert old_record["events"][0]["reason"] == "Superseded by a newer revision from the same source URI."


def test_pdf_tables_visuals_and_images_are_extracted_as_portable_artifacts(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "fixtures"
    fixture_dir.mkdir()
    pdf_path = fixture_dir / "power-sequencing-datasheet.pdf"
    image_path = fixture_dir / "debug-flowchart.png"
    make_table_and_flowchart_pdf(pdf_path)
    make_image(image_path)
    archive = DocumentationArchive(tmp_path / "archive")

    archive.ingest_path(pdf_path)
    archive.ingest_path(image_path)

    table_result = archive.search("POWER GOOD voltage current", source_types=["datasheet"], limit=1)[0]
    pdf_snapshot = tmp_path / "archive" / table_result["citation"]["snapshotPath"]
    extracted = pdf_snapshot.parent / "extracted"
    assert (extracted / "table_index.tsv").exists()
    assert "POWER GOOD" in (extracted / "tables" / "table-001.csv").read_text(encoding="utf-8")
    assert (extracted / "figure_index.tsv").exists()
    assert any(path.name.endswith(".png") for path in (extracted / "figures").iterdir())
    assert archive.search("flowcharts schematics visual layout", source_types=["datasheet"], limit=1)
    pdf_record = archive.get_document(table_result["documentId"])
    figure_artifact = next(artifact for artifact in pdf_record["artifacts"] if artifact["type"] == "figure")
    table_artifact = next(artifact for artifact in pdf_record["artifacts"] if artifact["type"] == "table")
    assert figure_artifact["path"].startswith("figures/")
    assert figure_artifact["mimeType"] == "image/png"
    assert figure_artifact["available"] is True
    assert table_artifact["alternatePaths"][0]["kind"] == "csv"
    assert table_artifact["nonEmptyCells"] >= 3
    assert table_artifact["totalCells"] >= table_artifact["nonEmptyCells"]
    served_figure = archive.document_artifact_file(table_result["documentId"], figure_artifact["path"])
    assert served_figure.media_type == "image/png"
    assert served_figure.path.is_file()
    with pytest.raises(ArchiveError):
        archive.document_artifact_file(table_result["documentId"], "../catalog.sqlite")

    client = TestClient(create_app(tmp_path / "archive"))
    artifact_response = client.get(f"/documents/{table_result['documentId']}/artifact", params={"path": figure_artifact["path"]})
    assert artifact_response.status_code == 200
    assert artifact_response.headers["content-type"].startswith("image/png")

    image_result = archive.search("320x160 flowcharts screenshots", source_types=["image"], limit=1)[0]
    image_snapshot = tmp_path / "archive" / image_result["citation"]["snapshotPath"]
    assert (image_snapshot.parent / "extracted" / "images" / "debug-flowchart.png").exists()
    image_record = archive.get_document(image_result["documentId"])
    assert image_record["artifacts"][0]["type"] == "image"
    assert image_record["artifacts"][0]["path"] == "images/debug-flowchart.png"


def test_schematic_pdf_page_creates_searchable_artifacts(tmp_path: Path) -> None:
    pdf_path = tmp_path / "buck-converter-schematic.pdf"
    make_schematic_pdf(pdf_path)
    archive = DocumentationArchive(tmp_path / "archive")

    document = archive.ingest_path(pdf_path, source_type="datasheet")[0]

    result = archive.search("Phase 1 analysis outputs R3 VDD schematic", source_types=["datasheet"], limit=1)[0]
    assert result["documentId"] == document.document_id
    assert result["locator"].startswith("schematic schematic-001 page 1 figure-001")
    snapshot = tmp_path / "archive" / result["citation"]["snapshotPath"]
    extracted = snapshot.parent / "extracted"
    assert (extracted / "schematic_index.tsv").exists()
    description_path = extracted / "schematics" / "schematic-001" / "description.md"
    analysis_path = extracted / "schematics" / "schematic-001" / "analysis.json"
    assert "Reference designators: R3, U1" in description_path.read_text(encoding="utf-8")
    analysis = json.loads(analysis_path.read_text(encoding="utf-8"))
    assert analysis["analysisOutputs"] == []
    assert analysis["referenceDesignators"] == ["R3", "U1"]
    assert {"GND", "VDD"}.issubset(set(analysis["labels"]))

    record = archive.get_document(document.document_id)
    schematic = next(artifact for artifact in record["artifacts"] if artifact["type"] == "schematic")
    assert schematic["kind"] == "schematic-description"
    assert schematic["imagePath"] == "figures/figure-001.png"
    assert schematic["descriptionPath"] == "schematics/schematic-001/description.md"
    assert schematic["jsonPath"] == "schematics/schematic-001/analysis.json"
    assert schematic["analysisOutputs"] == []
    assert archive.document_artifact_file(document.document_id, schematic["descriptionPath"]).media_type == "text/markdown"
    assert archive.document_artifact_file(document.document_id, schematic["imagePath"]).media_type == "image/png"


def test_schematic_image_creates_description_without_overclassifying_generic_images(tmp_path: Path) -> None:
    schematic_image = tmp_path / "r3-vdd-schematic.png"
    flowchart_image = tmp_path / "debug-flowchart.png"
    make_schematic_image(schematic_image)
    make_image(flowchart_image)
    archive = DocumentationArchive(tmp_path / "archive")

    schematic_document = archive.ingest_path(schematic_image)[0]
    flowchart_document = archive.ingest_path(flowchart_image)[0]

    result = archive.search("R3 VDD line-art schematic", source_types=["image"], limit=1)[0]
    assert result["documentId"] == schematic_document.document_id
    assert result["locator"] == "schematic schematic-001 image frame 1"
    schematic_record = archive.get_document(schematic_document.document_id)
    schematic_artifact = next(artifact for artifact in schematic_record["artifacts"] if artifact["type"] == "schematic")
    assert schematic_artifact["imagePath"] == "images/r3-vdd-schematic.png"
    assert schematic_artifact["referenceDesignators"] == ["R3"]
    assert "VDD" in schematic_artifact["labels"]

    flowchart_record = archive.get_document(flowchart_document.document_id)
    assert [artifact for artifact in flowchart_record["artifacts"] if artifact["type"] == "schematic"] == []
    assert all(not chunk["locator"].startswith("schematic ") for chunk in flowchart_record["chunks"])
    assert all("Analysis outputs are empty in Phase 1" not in chunk["text"] for chunk in flowchart_record["chunks"])
    flowchart_snapshot = tmp_path / "archive" / flowchart_record["snapshot_path"]
    assert not (flowchart_snapshot.parent / "extracted" / "schematic_index.tsv").exists()


def test_spreadsheet_workbook_extracts_searchable_sheet_artifacts(tmp_path: Path) -> None:
    workbook_path = tmp_path / "power-budget.xlsx"
    make_xlsx_workbook(workbook_path)
    archive = DocumentationArchive(tmp_path / "archive")

    document = archive.ingest_path(workbook_path)[0]

    assert document.source_type == "spreadsheet"
    result = archive.search("SPREADSHEET-REG-42 I2C gain", source_types=["spreadsheet"], limit=1)[0]
    assert result["documentId"] == document.document_id
    assert result["locator"] == "sheet Register Map range A1:C3"
    snapshot = tmp_path / "archive" / result["citation"]["snapshotPath"]
    extracted = snapshot.parent / "extracted"
    assert (extracted / "spreadsheet_index.tsv").exists()
    assert "SPREADSHEET-REG-42" in (extracted / "spreadsheets" / "sheet-002-Register_Map.csv").read_text(encoding="utf-8")

    record = archive.get_document(document.document_id)
    artifacts = [artifact for artifact in record["artifacts"] if artifact["type"] == "spreadsheet"]
    assert len(artifacts) == 2
    power_artifact = next(artifact for artifact in artifacts if artifact["sheet"] == "Power Budget")
    assert power_artifact["path"] == "spreadsheets/sheet-001-Power_Budget.md"
    assert power_artifact["mimeType"] == "text/markdown"
    assert power_artifact["available"] is True
    assert power_artifact["formulaCells"] == 1
    assert power_artifact["mergedRanges"] == ["A6:C6"]
    assert {alternate["kind"] for alternate in power_artifact["alternatePaths"]} == {"csv", "markdown", "json"}

    served_markdown = archive.document_artifact_file(document.document_id, power_artifact["path"])
    assert served_markdown.path.read_text(encoding="utf-8").startswith("# Sheet Power Budget")
    served_json = archive.document_artifact_file(document.document_id, power_artifact["jsonPath"])
    assert json.loads(served_json.path.read_text(encoding="utf-8"))["formulas"][0]["formula"] == "=SUM(C2:C3)"
    with pytest.raises(ArchiveError):
        archive.document_artifact_file(document.document_id, "spreadsheets/unknown.csv")


def test_directory_ingest_includes_xlsx_and_legacy_xls_workbooks(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "workbooks"
    fixture_dir.mkdir()
    make_xlsx_workbook(fixture_dir / "power-budget.xlsx")
    make_xls_workbook(fixture_dir / "legacy-registers.xls")
    archive = DocumentationArchive(tmp_path / "archive")

    documents = archive.ingest_path(fixture_dir)

    assert {document.source_type for document in documents} == {"spreadsheet"}
    assert archive.search("SPREADSHEET-REG-42", source_types=["spreadsheet"], limit=1)[0]["sourceType"] == "spreadsheet"
    legacy_result = archive.search("LEGACY-XLS-991 trim mode", source_types=["spreadsheet"], limit=1)[0]
    assert legacy_result["title"] == "legacy-registers.xls"
    legacy_record = archive.get_document(legacy_result["documentId"])
    legacy_artifact = next(artifact for artifact in legacy_record["artifacts"] if artifact["type"] == "spreadsheet")
    assert legacy_artifact["sheet"] == "Legacy XLS"
    assert legacy_artifact["csvPath"].endswith(".csv")


def test_fastapi_upload_autodetects_spreadsheet_source_type_and_serves_artifacts(tmp_path: Path) -> None:
    workbook_path = tmp_path / "upload-register-map.xlsx"
    make_xlsx_workbook(workbook_path)
    app = create_app(tmp_path / "archive")
    client = TestClient(app)

    response = client.post(
        "/ingest/upload",
        files={
            "file": (
                "upload-register-map.xlsx",
                workbook_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    document = response.json()["document"]
    assert document["sourceType"] == "spreadsheet"
    search = client.post("/search", json={"query": "SPREADSHEET-REG-42 I2C gain", "sourceTypes": ["spreadsheet"]})
    assert search.status_code == 200
    document_id = search.json()["results"][0]["documentId"]
    detail = client.get(f"/documents/{document_id}", params={"artifactLimit": 1})
    assert detail.status_code == 200
    artifact = detail.json()["document"]["artifacts"][0]
    assert detail.json()["document"]["artifactWindow"]["total"] == 2
    artifact_response = client.get(f"/documents/{document_id}/artifact", params={"path": artifact["path"]})
    assert artifact_response.status_code == 200
    assert "Sheet Power Budget" in artifact_response.text


def test_pdf_table_filter_rejects_single_cell_diagram_fragments() -> None:
    assert plausible_tables([[[None], ["EXPANDED TIME SCALE"]]]) == []

    tables = plausible_tables([[["Signal", "Min", "Max"], ["POWER GOOD", "1.7 V", "1.9 V"]]])
    assert len(tables) == 1
    assert tables[0].non_empty_cells == 6
    assert tables[0].total_cells == 6


def test_pdf_register_bit_tables_expand_merged_bit_cells() -> None:
    tables = plausible_tables([[
        ["BIT", "7", "6", "5", "4", "3", "2", "1", "0"],
        ["Field", "-", "mem_dt8_selu[6:0]", None, None, None, None, None, None],
        ["Reset", "-", "0x00", None, None, None, None, None, None],
        ["Access\nType", "-", "Write, Read", None, None, None, None, None, None],
    ]])

    assert len(tables) == 1
    table = tables[0]
    assert table.rows[1] == ["Field", "-", *(["mem_dt8_selu[6:0]"] * 7)]
    assert table.rows[3][0] == "Access Type"
    assert table.non_empty_cells == table.total_cells

    span = table_span_text("table-463", 145, table)
    assert "bits 6:0: Field mem_dt8_selu[6:0]; Reset 0x00; Access Type Write, Read" in span
    assert "| Field | - | mem_dt8_selu[6:0] | mem_dt8_selu[6:0] |" in span


def test_multiframe_images_preserve_every_frame_as_artifacts(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "fixtures"
    fixture_dir.mkdir()
    image_path = fixture_dir / "animated-flow.gif"
    make_multiframe_image(image_path)
    archive = DocumentationArchive(tmp_path / "archive")

    archive.ingest_path(image_path)

    result = archive.search("frames 2 frame artifacts", source_types=["image"], limit=1)[0]
    image_snapshot = tmp_path / "archive" / result["citation"]["snapshotPath"]
    extracted = image_snapshot.parent / "extracted"
    metadata = json.loads((extracted / "image_metadata.json").read_text(encoding="utf-8"))
    assert metadata["frames"] == 2
    assert metadata["artifacts"] == ["images/animated-flow-frame-0001.png", "images/animated-flow-frame-0002.png"]
    assert (extracted / "images" / "animated-flow-frame-0001.png").exists()
    assert (extracted / "images" / "animated-flow-frame-0002.png").exists()


def test_hybrid_search_preserves_strict_identifier_hits(tmp_path: Path) -> None:
    archive = DocumentationArchive(tmp_path / "archive")
    for index in range(8):
        archive.ingest_text(
            title=f"GMSL distractor {index}",
            text=("GMSL lane polarity override serializer deserializer table register bit " * 40).strip(),
            source_type="datasheet",
            uri=f"mock://distractor/{index}",
        )
    archive.ingest_text(
        title="Bulk mock needle",
        text=(
            "Bulk validation records. "
            "GMSL_SIM_088 lane polarity override is controlled by register 0x5A bit 3 in the simulated serializer."
        ),
        source_type="text",
        uri="mock://needle",
    )

    result = archive.search("GMSL_SIM_088 lane polarity override", limit=3, mode="hybrid")[0]

    assert result["title"] == "Bulk mock needle"


def test_rank_fusion_uses_lexical_relevance_to_protect_exact_matches() -> None:
    fused = reciprocal_rank_fusion(
        dense_scores={1: 0.9, 2: 0.8},
        lexical_scores={3: 8.0, 1: 1.0, 2: 0.5},
    )

    assert next(iter(fused)) == 3


def test_datasheet_source_type_does_not_force_non_pdf_through_pdf_extractor(tmp_path: Path) -> None:
    note = tmp_path / "power-rail-note.md"
    note.write_text("Power rail note\nThe CXRAIL-77 brownout threshold is 2.75 V.\n", encoding="utf-8")
    archive = DocumentationArchive(tmp_path / "archive")

    archive.ingest_path(note, source_type="datasheet")

    result = archive.search("CXRAIL-77 brownout threshold", limit=1)[0]
    assert result["sourceType"] == "datasheet"
    assert result["locator"] == "text"


def test_vendor_code_path_requires_review_and_generates_searchable_documentation(tmp_path: Path) -> None:
    source = tmp_path / "boot_driver.c"
    source_text = """
#define BOOT_DRIVER_MODE 3
static void apply_reset_delay(void) {
  hardware_write(0x40000010, BOOT_DRIVER_MODE);
  const char *raw_only = "RAW_ONLY_CODE_TOKEN";
}
void configure_boot_pin(void) {
  apply_reset_delay();
}
""".strip()
    source.write_text(source_text, encoding="utf-8")
    archive = DocumentationArchive(tmp_path / "archive")

    with pytest.raises(ArchiveError, match="acceptGeneratedCodeDocumentation=true"):
        archive.ingest_path(source)

    document = archive.ingest_path(source, accept_generated_code_documentation=True)[0]

    record = archive.get_document(document.document_id)
    assert record["source_type"] == "repo_code"
    assert record["snapshot_path"].endswith(".generated-code.md")
    assert archive.search("configure_boot_pin BOOT_DRIVER_MODE", source_types=["repo_code"], limit=1)[0]["documentId"] == document.document_id
    assert archive.search("RAW_ONLY_CODE_TOKEN", source_types=["repo_code"], limit=5, mode="lexical") == []

    artifacts = record["artifacts"]
    manifest_artifact = next(artifact for artifact in artifacts if artifact["type"] == "vendor_code" and artifact["kind"] == "manifest")
    assert manifest_artifact["rawSourceIndexed"] is False
    assert manifest_artifact["rawSourceRetained"] is False
    assert manifest_artifact["coveredFileCount"] == 1
    assert not any(artifact["type"] == "vendor_code" and artifact["kind"] == "source" for artifact in artifacts)

    manifest_file = archive.document_artifact_file(document.document_id, manifest_artifact["path"])
    manifest = json.loads(manifest_file.path.read_text(encoding="utf-8"))
    covered = manifest["coveredFiles"][0]
    assert covered["path"] == "boot_driver.c"
    assert covered["sha256"] == hashlib.sha256(source_text.encode("utf-8")).hexdigest()
    assert {"kind": "function", "name": "configure_boot_pin", "line": 6} in covered["symbols"]
    assert covered["artifactPath"] is None


def test_vendor_code_directory_mixes_docs_with_generated_repo_doc_and_rejects_unsupported_code(tmp_path: Path) -> None:
    fixture_dir = tmp_path / "vendor-drop"
    src_dir = fixture_dir / "src"
    src_dir.mkdir(parents=True)
    (fixture_dir / "README.md").write_text("Vendor guide says VENDOR-DOC-42 configures reset straps.\n", encoding="utf-8")
    (src_dir / "driver.py").write_text(
        "class ResetDriver:\n"
        "    def configure_reset_strap(self):\n"
        "        write_register(0x44, RESET_STRAP_MODE)\n",
        encoding="utf-8",
    )
    unsupported = src_dir / "rtl.sv"
    unsupported.write_text("module unsupported; endmodule\n", encoding="utf-8")
    archive = DocumentationArchive(tmp_path / "archive")

    with pytest.raises(ArchiveError, match="Unsupported code source files"):
        archive.ingest_path(fixture_dir, accept_generated_code_documentation=True)
    unsupported.unlink()

    with pytest.raises(ArchiveError, match="acceptGeneratedCodeDocumentation=true"):
        archive.ingest_path(fixture_dir)

    documents = archive.ingest_path(fixture_dir, accept_generated_code_documentation=True)

    assert sorted(document.source_type for document in documents) == ["readme", "repo_code"]
    assert archive.search("VENDOR-DOC-42 reset straps", source_types=["readme"], limit=1)[0]["sourceType"] == "readme"
    repo_result = archive.search("ResetDriver configure_reset_strap", source_types=["repo_code"], limit=1)[0]
    assert repo_result["sourceType"] == "repo_code"
    repo_record = archive.get_document(repo_result["documentId"])
    manifest = json.loads(archive.document_artifact_file(repo_result["documentId"], "vendor_code/code_manifest.json").path.read_text(encoding="utf-8"))
    assert manifest["coveredFiles"][0]["path"] == "src/driver.py"
    assert manifest["coveredFiles"][0]["parser"] == "python-ast"
    assert repo_record["collection"] == "vendor-drop"


def test_vendor_code_upload_url_and_text_policy(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    archive = DocumentationArchive(tmp_path / "archive")

    with pytest.raises(ArchiveError, match="Direct text repo_code ingest is not supported"):
        archive.ingest_text(title="Pasted source", text="def raw_source(): pass", source_type="repo_code")

    app = create_app(tmp_path / "api-archive")
    client = TestClient(app)
    rejected = client.post(
        "/ingest/upload",
        files={"file": ("sensor.ts", b"export function configureSensor() { return SENSOR_MODE; }\n", "text/typescript")},
    )
    assert rejected.status_code == 400
    assert "acceptGeneratedCodeDocumentation=true" in rejected.json()["detail"]

    accepted = client.post(
        "/ingest/upload",
        data={"acceptGeneratedCodeDocumentation": "true", "retainRawCodeArtifacts": "true"},
        files={"file": ("sensor.ts", b"export function configureSensor() { return SENSOR_MODE; }\n", "text/typescript")},
    )
    assert accepted.status_code == 200
    document = accepted.json()["document"]
    assert document["sourceType"] == "repo_code"
    uploaded_record = client.get(f"/documents/{document['documentId']}").json()["document"]
    assert any(artifact["type"] == "vendor_code" and artifact["kind"] == "source" for artifact in uploaded_record["artifacts"])

    class Response:
        headers = {"content-type": "text/plain"}
        url = "https://vendor.example/driver.py"

    monkeypatch.setattr(
        archive_module,
        "fetch_url_bytes",
        lambda url, limit: (Response(), b"def configure_url_driver():\n    return URL_DRIVER_MODE\n"),
    )
    url_doc = archive.ingest_url("https://vendor.example/driver.py", accept_generated_code_documentation=True)
    assert archive.search("configure_url_driver URL_DRIVER_MODE", source_types=["repo_code"], limit=1)[0]["documentId"] == url_doc.document_id


def test_fastapi_upload_ingests_documentation_file(tmp_path: Path) -> None:
    app = create_app(tmp_path / "archive")
    client = TestClient(app)

    response = client.post(
        "/ingest/upload",
        data={"sourceType": "readme", "collection": "uploads"},
        files={"file": ("uploaded-note.md", b"Uploaded note says UPLOAD-NEEDLE-41 uses hybrid recall.\n", "text/markdown")},
    )

    assert response.status_code == 200
    assert response.json()["document"]["sourceType"] == "readme"
    search = client.post("/search", json={"query": "UPLOAD-NEEDLE-41 hybrid recall", "collection": "uploads"})
    assert search.status_code == 200
    result = search.json()["results"][0]
    assert result["sourceType"] == "readme"
    assert result["citation"]["snapshotPath"].endswith("/uploaded-note.md")


def test_fastapi_path_ingest_rejects_relative_paths_from_service_cwd(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service_cwd = tmp_path / "service"
    workspace = tmp_path / "workspace"
    service_cwd.mkdir()
    workspace.mkdir()
    note = workspace / "relative-note.md"
    note.write_text("Relative note says RELATIVE-PATH-77 resolves through CloudX.\n", encoding="utf-8")
    monkeypatch.chdir(service_cwd)
    app = create_app(tmp_path / "archive")
    client = TestClient(app)

    rejected = client.post("/ingest/path", json={"path": "relative-note.md"})

    assert rejected.status_code == 400
    assert "absolute path" in rejected.json()["detail"]

    accepted = client.post("/ingest/path", json={"path": str(note), "sourceType": "readme"})
    assert accepted.status_code == 200
    search = client.post("/search", json={"query": "RELATIVE-PATH-77 CloudX", "sourceTypes": ["readme"]})
    assert search.status_code == 200
    assert search.json()["results"][0]["sourceType"] == "readme"


def test_archive_path_ingest_rejects_relative_paths(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    service_cwd = tmp_path / "service"
    workspace = tmp_path / "workspace"
    service_cwd.mkdir()
    workspace.mkdir()
    (workspace / "relative-note.md").write_text("Relative note should not resolve from service cwd.\n", encoding="utf-8")
    monkeypatch.chdir(service_cwd)
    archive = DocumentationArchive(tmp_path / "archive")

    with pytest.raises(ArchiveError, match="absolute path"):
        archive.ingest_path("../workspace/relative-note.md")


def test_url_fetch_rejects_unsupported_schemes_and_oversized_downloads() -> None:
    with pytest.raises(ArchiveError, match="only http and https"):
        fetch_url_bytes("file:///etc/passwd", 1024)

    server = HtmlFixtureServer("0123456789abcdef")
    server.start()
    try:
        with pytest.raises(ArchiveError, match="exceeds the maximum size"):
            fetch_url_bytes(server.url, 8)
    finally:
        server.stop()


def test_concurrent_mutations_serialize_index_rebuilds(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    active_writes = 0
    max_active_writes = 0
    state_lock = threading.Lock()

    class SlowIdMapIndex:
        def __init__(self, **_kwargs: object):
            pass

        def add_with_ids(self, *_args: object) -> None:
            pass

        def write(self, path: str) -> None:
            nonlocal active_writes, max_active_writes
            with state_lock:
                active_writes += 1
                max_active_writes = max(max_active_writes, active_writes)
            try:
                time.sleep(0.05)
                Path(path).write_bytes(b"fake-index")
            finally:
                with state_lock:
                    active_writes -= 1

    monkeypatch.setattr(archive_module, "IdMapIndex", SlowIdMapIndex)
    archive = DocumentationArchive(tmp_path / "archive")

    with ThreadPoolExecutor(max_workers=2) as executor:
        futures = [
            executor.submit(
                archive.ingest_text,
                title=f"Concurrent document {index}",
                text=f"Concurrent ingest {index} says LOCK-SERIAL-{index}.",
                uri=f"mock://concurrent/{index}",
            )
            for index in range(2)
        ]
        documents = [future.result(timeout=5) for future in futures]

    assert max_active_writes == 1
    assert {document.title for document in documents} == {"Concurrent document 0", "Concurrent document 1"}
    assert len(archive.list_documents()) == 2


def test_cli_help_documents_service_options() -> None:
    result = subprocess.run(
        [sys.executable, "-m", "cloudx_documentation_indexer.main", "--help"],
        check=True,
        capture_output=True,
        text=True,
    )

    assert "--archive-root" in result.stdout
    assert "--host" in result.stdout
    assert "--port" in result.stdout


def make_pdf(path: Path, lines: list[str]) -> None:
    pdf = canvas.Canvas(str(path))
    y = 760
    for line in lines:
        pdf.drawString(72, y, line)
        y -= 22
    pdf.save()


def make_table_and_flowchart_pdf(path: Path) -> None:
    pdf = canvas.Canvas(str(path))
    pdf.drawString(72, 760, "Power Sequencing Datasheet")
    pdf.drawString(72, 736, "The POWER GOOD signal must be high before enabling the load switch.")
    x0, y0 = 72, 680
    col_widths = [130, 120, 120]
    row_height = 24
    rows = [
        ["Signal", "Voltage", "Current"],
        ["POWER GOOD", "3.3 V", "12 mA"],
        ["LOAD ENABLE", "1.8 V", "4 mA"],
    ]
    for row_index, row in enumerate(rows):
        y = y0 - row_index * row_height
        x = x0
        for col_index, cell in enumerate(row):
            pdf.rect(x, y - row_height, col_widths[col_index], row_height)
            pdf.drawString(x + 5, y - 16, cell)
            x += col_widths[col_index]
    pdf.drawString(72, 560, "Power sequencing flowchart")
    pdf.rect(72, 500, 120, 36)
    pdf.drawString(92, 514, "RESET LOW")
    pdf.line(192, 518, 260, 518)
    pdf.line(252, 524, 260, 518)
    pdf.line(252, 512, 260, 518)
    pdf.rect(260, 500, 130, 36)
    pdf.drawString(282, 514, "ENABLE RAIL")
    pdf.save()


def make_schematic_pdf(path: Path) -> None:
    pdf = canvas.Canvas(str(path))
    pdf.drawString(72, 760, "Buck Converter Schematic")
    pdf.drawString(72, 736, "U1 drives R3 from VDD to GND in the feedback circuit.")
    pdf.rect(180, 610, 120, 80)
    pdf.drawString(216, 650, "U1")
    pdf.drawString(206, 632, "REGULATOR")
    pdf.line(72, 650, 180, 650)
    pdf.drawString(78, 662, "VDD")
    pdf.line(300, 650, 430, 650)
    pdf.rect(430, 630, 50, 40)
    pdf.drawString(444, 645, "R3")
    pdf.line(480, 650, 520, 650)
    pdf.line(520, 650, 520, 590)
    pdf.line(500, 590, 540, 590)
    pdf.line(506, 582, 534, 582)
    pdf.line(514, 574, 526, 574)
    pdf.drawString(548, 586, "GND")
    pdf.save()


def make_image(path: Path) -> None:
    image = Image.new("RGB", (320, 160), "white")
    draw = ImageDraw.Draw(image)
    draw.rectangle((24, 48, 128, 96), outline="black", width=3)
    draw.text((42, 63), "START", fill="black")
    draw.line((128, 72, 200, 72), fill="black", width=3)
    draw.rectangle((200, 48, 296, 96), outline="black", width=3)
    draw.text((218, 63), "DONE", fill="black")
    image.save(path)


def make_schematic_image(path: Path) -> None:
    image = Image.new("RGB", (360, 180), "white")
    draw = ImageDraw.Draw(image)
    draw.rectangle((142, 58, 218, 118), outline="black", width=3)
    draw.text((168, 78), "U1", fill="black")
    draw.line((40, 88, 142, 88), fill="black", width=3)
    draw.text((48, 64), "VDD", fill="black")
    draw.rectangle((236, 72, 286, 104), outline="black", width=3)
    draw.text((250, 80), "R3", fill="black")
    draw.line((218, 88, 236, 88), fill="black", width=3)
    draw.line((286, 88, 326, 88), fill="black", width=3)
    draw.line((326, 88, 326, 132), fill="black", width=3)
    draw.line((306, 132, 346, 132), fill="black", width=3)
    draw.line((313, 140, 339, 140), fill="black", width=3)
    draw.line((321, 148, 331, 148), fill="black", width=3)
    draw.text((292, 152), "GND", fill="black")
    image.save(path)


def make_multiframe_image(path: Path) -> None:
    first = Image.new("RGB", (160, 90), "white")
    draw = ImageDraw.Draw(first)
    draw.rectangle((16, 24, 72, 64), outline="black", width=3)
    draw.text((28, 38), "ONE", fill="black")
    second = Image.new("RGB", (160, 90), "white")
    draw = ImageDraw.Draw(second)
    draw.rectangle((88, 24, 144, 64), outline="black", width=3)
    draw.text((99, 38), "TWO", fill="black")
    first.save(path, save_all=True, append_images=[second], duration=100, loop=0)


def make_xlsx_workbook(path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    power = workbook.active
    power.title = "Power Budget"
    power.append(["Component", "Rail", "Current_mA"])
    power.append(["MCU", "3V3", 45])
    power.append(["Sensor", "1V8", 12])
    power.append(["Total", "", "=SUM(C2:C3)"])
    power["A6"] = "Notes: merged range captures board bring-up context"
    power.merge_cells("A6:C6")
    registers = workbook.create_sheet("Register Map")
    registers.append(["Register", "Address", "Purpose"])
    registers.append(["SPREADSHEET-REG-42", "0x2A", "I2C gain trim"])
    registers.append(["BOOT_DELAY", "0x2B", "reset delay cycles"])
    workbook.save(path)


def make_xls_workbook(path: Path) -> None:
    import xlwt

    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Legacy XLS")
    rows = [
        ["Name", "Value", "Purpose"],
        ["LEGACY-XLS-991", 17, "trim mode"],
        ["LEGACY-XLS-992", 23, "reset mode"],
    ]
    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            sheet.write(row_index, column_index, value)
    workbook.save(str(path))


def stub_youtube_media(monkeypatch: pytest.MonkeyPatch, transcripts: dict[str, str], descriptions: dict[str, str | None] | None = None) -> None:
    monkeypatch.setattr(archive_module, "fetch_youtube_transcript", lambda url: transcripts[url])
    monkeypatch.setattr(
        archive_module,
        "transcribe_youtube_video",
        lambda url, _metadata, progress=None: [
            archive_module.TranscriptSegment(0.0, 2.0, transcripts[url]),
        ],
    )

    def metadata(url: str) -> archive_module.YouTubeVideoMetadata:
        video_id = url.rsplit("=", 1)[-1]
        return archive_module.YouTubeVideoMetadata(
            title=f"Video {video_id}",
            webpage_url=url,
            stream_url=f"mock://stream/{video_id}",
            http_headers={"User-Agent": "cloudx-test"},
            duration=1234,
            uploader="CloudX Test",
            upload_date="20260607",
            description=descriptions.get(url, f"Metadata for {video_id} includes allocator pressure slides.") if descriptions is not None else f"Metadata for {video_id} includes allocator pressure slides.",
            thumbnail=f"https://img.youtube.com/vi/{video_id}/0.jpg",
            tags=["memory", "slides"],
            chapters=[{"title": "Intro", "start_time": 0, "end_time": 60}],
        )

    def keyframes(_metadata: archive_module.YouTubeVideoMetadata, artifact_dir: Path, *, transcript_segments=None, progress=None) -> list[dict[str, object]]:
        frames_dir = artifact_dir / "media" / "keyframes"
        frames_dir.mkdir(parents=True, exist_ok=True)
        for index in range(1, 3):
            Image.new("RGB", (64, 36), "white").save(frames_dir / f"frame-{index:06d}.jpg")
        frames = [
            {"offsetSeconds": 0, "path": "media/keyframes/frame-000001.jpg", "reason": "segment-start", "transcriptStartSeconds": 0.0, "transcriptEndSeconds": 2.0},
            {"offsetSeconds": 2, "path": "media/keyframes/frame-000002.jpg", "reason": "visual-change", "changeScore": 0.5, "transcriptStartSeconds": 0.0, "transcriptEndSeconds": 2.0},
        ]
        media_dir = artifact_dir / "media"
        media_dir.mkdir(parents=True, exist_ok=True)
        archive_module.write_keyframe_index(media_dir / "keyframes.tsv", frames)
        archive_module.write_visual_sampling_manifest(
            media_dir / "visual_sampling.json",
            metadata=_metadata,
            profile=archive_module.VideoVisualProfile(),
            scanned_frames=4,
            selected_frames=2,
            elapsed_seconds=0.1,
        )
        (media_dir / "youtube_metadata.json").write_text(json.dumps(archive_module.youtube_metadata_json(_metadata), indent=2) + "\n", encoding="utf-8")
        if _metadata.description:
            (media_dir / "description.txt").write_text(_metadata.description + "\n", encoding="utf-8")
        return frames

    monkeypatch.setattr(archive_module, "extract_youtube_video_metadata", metadata)
    monkeypatch.setattr(archive_module, "capture_youtube_keyframes", keyframes)


def dense_collision_query(reference_tokens: list[str]) -> str:
    return " ".join(dense_collision_token(token) for token in reference_tokens)


def dense_collision_token(reference_token: str) -> str:
    target = dense_token_signature(reference_token)
    for attempt in range(100_000):
        candidate = f"densecollision{attempt}_{reference_token}"
        if dense_token_signature(candidate) == target:
            return candidate
    raise AssertionError(f"No dense collision found for {reference_token}.")


def dense_token_signature(token: str) -> tuple[int, bool]:
    digest = hashlib.sha256(token.encode("utf-8")).digest()
    return int.from_bytes(digest[:4], "big") % EMBEDDING_DIM, bool(digest[4] & 1)


class HtmlFixtureServer:
    def __init__(self, html: str):
        self.html = html.encode("utf-8")
        self.server: ThreadingHTTPServer | None = None
        self.thread: threading.Thread | None = None
        self.url = ""

    def start(self) -> None:
        html = self.html

        class Handler(BaseHTTPRequestHandler):
            def do_GET(self):  # noqa: N802
                self.send_response(200)
                self.send_header("content-type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(html)

            def log_message(self, format, *args):  # noqa: A002
                return

        self.server = ThreadingHTTPServer(("127.0.0.1", 0), Handler)
        self.url = f"http://127.0.0.1:{self.server.server_port}/thermal.html"
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)
        self.thread.start()

    def stop(self) -> None:
        if self.server:
            self.server.shutdown()
            self.server.server_close()
        if self.thread:
            self.thread.join(timeout=5)
