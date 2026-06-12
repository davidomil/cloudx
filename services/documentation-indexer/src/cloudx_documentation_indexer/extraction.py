from __future__ import annotations

import csv
from datetime import date, datetime, time
import io
import json
import math
import mimetypes
import re
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from bs4 import BeautifulSoup
from PIL import Image, ImageFilter, ImageSequence
import pdfplumber
import pypdfium2 as pdfium


PDF_SUFFIXES = {".pdf"}
HTML_SUFFIXES = {".html", ".htm"}
IMAGE_SUFFIXES = {".bmp", ".gif", ".jpeg", ".jpg", ".png", ".tif", ".tiff", ".webp"}
SPREADSHEET_SUFFIXES = {".xls", ".xlsx", ".xlsm", ".xlsb", ".ods", ".ots"}
SPREADSHEET_CONTENT_TYPES = {
    "application/vnd.ms-excel",
    "application/vnd.ms-excel.sheet.binary.macroenabled.12",
    "application/vnd.ms-excel.sheet.macroenabled.12",
    "application/vnd.oasis.opendocument.spreadsheet",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
SCHEMATIC_SCHEMA_VERSION = 1
SCHEMATIC_TERMS = {"schematic", "circuit", "netlist", "reference designator", "power rail"}
REFERENCE_DESIGNATOR_RE = re.compile(r"\b(?:R|C|L|U|J|P|Q|D|TP|FB|Y|X|K|F|SW|RN)\d+[A-Za-z]?\b")
NET_LABEL_RE = re.compile(r"\b(?:VCC|VDD|VSS|GND|AGND|DGND|VIN|VOUT|SDA|SCL|MISO|MOSI|RESET|ENABLE|EN|BOOT|INT|CLK|TX|RX)\b")
TEXT_SUFFIXES = {
    ".adoc",
    ".asciidoc",
    ".c",
    ".cpp",
    ".css",
    ".csv",
    ".h",
    ".hpp",
    ".js",
    ".json",
    ".md",
    ".py",
    ".rs",
    ".srt",
    ".text",
    ".tex",
    ".ts",
    ".tsx",
    ".txt",
    ".vtt",
    ".xml",
    ".yaml",
    ".yml",
}
SUPPORTED_FILE_SUFFIXES = PDF_SUFFIXES | HTML_SUFFIXES | IMAGE_SUFFIXES | SPREADSHEET_SUFFIXES | TEXT_SUFFIXES


@dataclass(frozen=True)
class ExtractedSpan:
    text: str
    locator: str


@dataclass(frozen=True)
class ExtractedTable:
    rows: list[list[str]]
    non_empty_cells: int
    total_cells: int


@dataclass(frozen=True)
class ExtractedSpreadsheetSheet:
    id: str
    name: str
    rows: list[list[str]]
    range_ref: str
    formulas: list[dict[str, str]]
    merged_ranges: list[str]


@dataclass(frozen=True)
class SchematicArtifact:
    id: str
    source: str
    locator: str
    image: str
    width: int
    height: int
    reasons: list[str]
    reference_designators: list[str]
    labels: list[str]
    connection_cues: list[str]
    metrics: dict[str, int | float]


def extract_file(path: Path, content: bytes, source_type: str, artifact_dir: Path | None = None) -> list[ExtractedSpan]:
    return extract_bytes(content, path.name, source_type, mimetypes.guess_type(path.name)[0], artifact_dir)


def extract_bytes(
    content: bytes,
    name: str,
    source_type: str,
    content_type: str | None,
    artifact_dir: Path | None = None,
) -> list[ExtractedSpan]:
    suffix = source_suffix(name)
    normalized_type = normalized_content_type(content_type)
    if suffix in PDF_SUFFIXES or normalized_type == "application/pdf" or looks_like_pdf(content):
        return PdfExtractionPipeline(artifact_dir).extract(content, name)
    if source_type == "spreadsheet" or suffix in SPREADSHEET_SUFFIXES or normalized_type in SPREADSHEET_CONTENT_TYPES:
        return SpreadsheetExtractionPipeline(artifact_dir).extract(content, name, normalized_type)
    if source_type == "image" or suffix in IMAGE_SUFFIXES or normalized_type.startswith("image/"):
        return ImageExtractionPipeline(artifact_dir).extract(content, name)
    if source_type == "website" or suffix in HTML_SUFFIXES or normalized_type.startswith("text/html"):
        return [ExtractedSpan(extract_html(content), "html")]
    return [ExtractedSpan(decode_text(content), "text")]


class PdfExtractionPipeline:
    def __init__(self, artifact_dir: Path | None = None):
        self.artifact_dir = artifact_dir

    def extract(self, content: bytes, name: str = "source.pdf") -> list[ExtractedSpan]:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as temp_file:
            temp_file.write(content)
            temp_path = Path(temp_file.name)
        try:
            return self._extract_from_path(temp_path, name)
        finally:
            temp_path.unlink(missing_ok=True)

    def _extract_from_path(self, pdf_path: Path, name: str) -> list[ExtractedSpan]:
        spans: list[ExtractedSpan] = []
        table_records: list[dict[str, str | int]] = []
        figure_records: list[dict[str, str | int]] = []
        schematic_records: list[dict[str, Any]] = []
        artifact_root = prepare_artifact_dir(self.artifact_dir)

        with pdfplumber.open(pdf_path) as pdf:
            pdfium_doc = pdfium.PdfDocument(str(pdf_path)) if artifact_root else None
            try:
                for page_number, page in enumerate(pdf.pages, start=1):
                    text = (page.extract_text() or "").strip()
                    if text:
                        spans.append(ExtractedSpan(text, f"page {page_number}"))

                    page_tables = plausible_tables(page.extract_tables() or [])
                    for table_number, table in enumerate(page_tables, start=1):
                        table_id = f"table-{len(table_records) + 1:03d}"
                        table_records.append(write_table_artifacts(artifact_root, table_id, page_number, table))
                        spans.append(ExtractedSpan(table_span_text(table_id, page_number, table), f"page {page_number} {table_id}"))

                    visual_summary = page_visual_summary(page)
                    if visual_summary and artifact_root and pdfium_doc:
                        figure_id = f"figure-{len(figure_records) + 1:03d}"
                        figure_path = render_page_artifact(pdfium_doc, page_number, artifact_root / "figures" / f"{figure_id}.png")
                        relative_figure_path = figure_path.relative_to(artifact_root).as_posix()
                        figure_records.append({
                            "id": figure_id,
                            "page": page_number,
                            "kind": "page-render",
                            "path": relative_figure_path,
                            **visual_summary,
                        })
                        spans.append(ExtractedSpan(visual_span_text(figure_id, page_number, visual_summary), f"page {page_number} {figure_id}"))
                        schematic = schematic_artifact_from_pdf_page(
                            schematic_id=f"schematic-{len(schematic_records) + 1:03d}",
                            filename=name,
                            page_number=page_number,
                            figure_id=figure_id,
                            image_path=relative_figure_path,
                            image_file=figure_path,
                            page_text=text,
                            visual_summary=visual_summary,
                        )
                        if schematic:
                            record = write_schematic_artifact(artifact_root, schematic)
                            schematic_records.append(record)
                            spans.append(ExtractedSpan(schematic_span_text(schematic, record), f"schematic {schematic.id} page {page_number} {figure_id}"))
            finally:
                if pdfium_doc is not None:
                    pdfium_doc.close()

        if artifact_root:
            write_index_files(artifact_root, table_records, figure_records, schematic_records)
        return spans


class ImageExtractionPipeline:
    def __init__(self, artifact_dir: Path | None = None):
        self.artifact_dir = artifact_dir

    def extract(self, content: bytes, name: str) -> list[ExtractedSpan]:
        spans: list[ExtractedSpan] = []
        schematic_records: list[dict[str, Any]] = []
        artifact_root = prepare_artifact_dir(self.artifact_dir)
        with Image.open(io.BytesIO(content)) as image:
            image.load()
            frames = frame_count(image)
            metadata: dict[str, Any] = {
                "filename": name,
                "format": image.format or Path(name).suffix.lstrip(".").upper() or "IMAGE",
                "width": image.width,
                "height": image.height,
                "mode": image.mode,
                "frames": frames,
            }
            if artifact_root:
                images_dir = artifact_root / "images"
                images_dir.mkdir(parents=True, exist_ok=True)
                frame_paths: list[str] = []
                for frame_index, frame in enumerate(ImageSequence.Iterator(image), start=1):
                    output_name = f"{safe_stem(name)}.png" if frames == 1 else f"{safe_stem(name)}-frame-{frame_index:04d}.png"
                    normalized_path = images_dir / output_name
                    normalized_frame = frame.convert("RGBA")
                    normalized_frame.save(normalized_path)
                    relative_frame_path = normalized_path.relative_to(artifact_root).as_posix()
                    frame_paths.append(relative_frame_path)
                    schematic = schematic_artifact_from_image(
                        schematic_id=f"schematic-{len(schematic_records) + 1:03d}",
                        filename=name,
                        frame_index=frame_index,
                        frames=frames,
                        image_path=relative_frame_path,
                        image=normalized_frame,
                    )
                    if schematic:
                        record = write_schematic_artifact(artifact_root, schematic)
                        schematic_records.append(record)
                        spans.append(ExtractedSpan(schematic_span_text(schematic, record), f"schematic {schematic.id} image frame {frame_index}"))
                metadata["artifact"] = frame_paths[0] if frame_paths else ""
                metadata["artifacts"] = frame_paths
                (artifact_root / "image_metadata.json").write_text(json.dumps(metadata, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        if artifact_root and schematic_records:
            write_schematic_index(artifact_root, schematic_records)
        return [ExtractedSpan(image_span_text(metadata), "image"), *spans]


class SpreadsheetExtractionPipeline:
    def __init__(self, artifact_dir: Path | None = None):
        self.artifact_dir = artifact_dir

    def extract(self, content: bytes, name: str, content_type: str | None = None) -> list[ExtractedSpan]:
        suffix = source_suffix(name)
        artifact_root = prepare_artifact_dir(self.artifact_dir)
        if suffix in {".xlsx", ".xlsm"} or content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            sheets = extract_ooxml_workbook(content, suffix)
        else:
            sheets = extract_pandas_workbook(content, suffix)
        records = write_spreadsheet_artifacts(artifact_root, sheets)
        return [
            ExtractedSpan(
                spreadsheet_sheet_span_text(sheet, records.get(sheet.id)),
                f"sheet {sheet.name} range {sheet.range_ref}",
            )
            for sheet in sheets
        ]


def prepare_artifact_dir(artifact_dir: Path | None) -> Path | None:
    if artifact_dir is None:
        return None
    shutil.rmtree(artifact_dir, ignore_errors=True)
    artifact_dir.mkdir(parents=True, exist_ok=True)
    return artifact_dir


def plausible_tables(tables: list[list[list[str | None]]]) -> list[ExtractedTable]:
    extracted: list[ExtractedTable] = []
    for table in tables:
        rows = clean_table_rows(table)
        if not rows:
            continue
        column_count = max((len(row) for row in rows), default=0)
        non_empty_cells = sum(1 for row in rows for cell in row if cell)
        total_cells = max(1, len(rows) * max(1, column_count))
        if is_plausible_table(rows, column_count, non_empty_cells):
            extracted.append(ExtractedTable(rows, non_empty_cells, total_cells))
    return extracted


def is_plausible_table(rows: list[list[str]], column_count: int, non_empty_cells: int) -> bool:
    return len(rows) >= 2 and column_count >= 2 and non_empty_cells >= 3


def clean_table_rows(table: list[list[str | None]]) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in table:
        cleaned = [clean_table_cell(cell) for cell in row]
        if any(cleaned):
            rows.append(cleaned)
    return expand_register_bit_table(rows)


def clean_table_cell(cell: str | None) -> str:
    return " ".join((cell or "").split())


def expand_register_bit_table(rows: list[list[str]]) -> list[list[str]]:
    if not rows or not is_register_bit_header(rows[0]):
        return rows
    width = len(rows[0])
    expanded = [rows[0]]
    for row in rows[1:]:
        padded = row + [""] * (width - len(row))
        if normalized_table_label(padded[0]) in {"field", "reset", "access", "access type"}:
            expanded.append(expand_merged_register_cells(padded))
        else:
            expanded.append(padded)
    return expanded


def expand_merged_register_cells(row: list[str]) -> list[str]:
    expanded = row[:]
    previous = ""
    for index in range(1, len(expanded)):
        if expanded[index]:
            previous = expanded[index]
        elif previous:
            expanded[index] = previous
    return expanded


def is_register_bit_header(row: list[str]) -> bool:
    if len(row) < 3 or normalized_table_label(row[0]) != "bit":
        return False
    try:
        bits = [int(cell, 10) for cell in row[1:]]
    except ValueError:
        return False
    return bits == list(range(bits[0], bits[-1] - 1, -1))


def normalized_table_label(value: str) -> str:
    return " ".join(value.lower().split())


def write_table_artifacts(artifact_root: Path | None, table_id: str, page_number: int, table: ExtractedTable) -> dict[str, str | int]:
    record: dict[str, str | int] = {
        "id": table_id,
        "page": page_number,
        "rows": len(table.rows),
        "columns": max((len(row) for row in table.rows), default=0),
        "non_empty_cells": table.non_empty_cells,
        "total_cells": table.total_cells,
    }
    if not artifact_root:
        return record
    tables_dir = artifact_root / "tables"
    tables_dir.mkdir(parents=True, exist_ok=True)
    csv_path = tables_dir / f"{table_id}.csv"
    md_path = tables_dir / f"{table_id}.md"
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(table.rows)
    md_path.write_text(table_markdown_document(table), encoding="utf-8")
    record["csv"] = csv_path.relative_to(artifact_root).as_posix()
    record["markdown"] = md_path.relative_to(artifact_root).as_posix()
    return record


def extract_ooxml_workbook(content: bytes, suffix: str) -> list[ExtractedSpreadsheetSheet]:
    from openpyxl import load_workbook

    keep_vba = suffix == ".xlsm"
    formulas_workbook = load_workbook(io.BytesIO(content), data_only=False, read_only=False, keep_vba=keep_vba)
    values_workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=False, keep_vba=keep_vba)
    sheets: list[ExtractedSpreadsheetSheet] = []
    for index, formula_sheet in enumerate(formulas_workbook.worksheets, start=1):
        value_sheet = values_workbook[formula_sheet.title]
        raw_rows: list[list[str]] = []
        formulas: list[dict[str, str]] = []
        for row_index in range(1, formula_sheet.max_row + 1):
            row: list[str] = []
            for column_index in range(1, formula_sheet.max_column + 1):
                formula_cell = formula_sheet.cell(row_index, column_index)
                value_cell = value_sheet.cell(row_index, column_index)
                row.append(spreadsheet_cell_text(formula_cell.value, value_cell.value))
                if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                    formulas.append(
                        {
                            "cell": spreadsheet_cell_ref(row_index, column_index),
                            "formula": formula_cell.value,
                            "cached_value": spreadsheet_value_text(value_cell.value),
                        }
                    )
            raw_rows.append(row)
        rows, range_ref = trim_spreadsheet_grid(raw_rows)
        sheets.append(
            ExtractedSpreadsheetSheet(
                id=f"sheet-{index:03d}",
                name=formula_sheet.title,
                rows=rows,
                range_ref=range_ref,
                formulas=[{key: value for key, value in formula.items() if value} for formula in formulas],
                merged_ranges=[str(merged_range) for merged_range in formula_sheet.merged_cells.ranges],
            )
        )
    return sheets


def extract_pandas_workbook(content: bytes, suffix: str) -> list[ExtractedSpreadsheetSheet]:
    import pandas as pd

    workbook = pd.ExcelFile(io.BytesIO(content), engine=pandas_engine_for_suffix(suffix))
    sheets: list[ExtractedSpreadsheetSheet] = []
    for index, sheet_name in enumerate(workbook.sheet_names, start=1):
        frame = workbook.parse(sheet_name=sheet_name, header=None, dtype=object, keep_default_na=False, na_filter=False)
        raw_rows = [[spreadsheet_value_text(value) for value in row] for row in frame.itertuples(index=False, name=None)]
        rows, range_ref = trim_spreadsheet_grid(raw_rows)
        sheets.append(
            ExtractedSpreadsheetSheet(
                id=f"sheet-{index:03d}",
                name=str(sheet_name),
                rows=rows,
                range_ref=range_ref,
                formulas=[],
                merged_ranges=[],
            )
        )
    return sheets


def pandas_engine_for_suffix(suffix: str) -> str:
    engines = {
        ".xls": "xlrd",
        ".xlsb": "pyxlsb",
        ".ods": "odf",
        ".ots": "odf",
        ".xlsx": "openpyxl",
        ".xlsm": "openpyxl",
    }
    return engines.get(suffix, "openpyxl")


def write_spreadsheet_artifacts(artifact_root: Path | None, sheets: list[ExtractedSpreadsheetSheet]) -> dict[str, dict[str, str | int]]:
    records: dict[str, dict[str, str | int]] = {}
    for sheet in sheets:
        record = spreadsheet_artifact_record(sheet)
        records[sheet.id] = record
        if artifact_root:
            spreadsheets_dir = artifact_root / "spreadsheets"
            spreadsheets_dir.mkdir(parents=True, exist_ok=True)
            base_name = f"{sheet.id}-{safe_stem(sheet.name)}"
            csv_path = spreadsheets_dir / f"{base_name}.csv"
            md_path = spreadsheets_dir / f"{base_name}.md"
            json_path = spreadsheets_dir / f"{base_name}.json"
            with csv_path.open("w", encoding="utf-8", newline="") as handle:
                writer = csv.writer(handle)
                writer.writerows(sheet.rows)
            md_path.write_text(spreadsheet_markdown_document(sheet), encoding="utf-8")
            json_path.write_text(json.dumps(spreadsheet_json_document(sheet), indent=2, sort_keys=True) + "\n", encoding="utf-8")
            record["csv"] = csv_path.relative_to(artifact_root).as_posix()
            record["markdown"] = md_path.relative_to(artifact_root).as_posix()
            record["json"] = json_path.relative_to(artifact_root).as_posix()
    if artifact_root and records:
        write_tsv(
            artifact_root / "spreadsheet_index.tsv",
            list(records.values()),
            ["id", "sheet", "range", "rows", "columns", "non_empty_cells", "formula_cells", "merged_ranges", "csv", "markdown", "json"],
        )
    return records


def spreadsheet_artifact_record(sheet: ExtractedSpreadsheetSheet) -> dict[str, str | int]:
    return {
        "id": sheet.id,
        "sheet": sheet.name,
        "range": sheet.range_ref,
        "rows": len(sheet.rows),
        "columns": max((len(row) for row in sheet.rows), default=0),
        "non_empty_cells": sum(1 for row in sheet.rows for cell in row if cell),
        "formula_cells": len(sheet.formulas),
        "merged_ranges": ", ".join(sheet.merged_ranges),
    }


def spreadsheet_json_document(sheet: ExtractedSpreadsheetSheet) -> dict[str, Any]:
    return {
        "id": sheet.id,
        "sheet": sheet.name,
        "range": sheet.range_ref,
        "rows": sheet.rows,
        "formulas": sheet.formulas,
        "mergedRanges": sheet.merged_ranges,
    }


def spreadsheet_markdown_document(sheet: ExtractedSpreadsheetSheet) -> str:
    lines = [
        f"# Sheet {sheet.name}",
        "",
        f"Range: {sheet.range_ref}",
        f"Rows: {len(sheet.rows)}",
        f"Columns: {max((len(row) for row in sheet.rows), default=0)}",
        f"Formula cells: {len(sheet.formulas)}",
    ]
    if sheet.merged_ranges:
        lines.append(f"Merged ranges: {', '.join(sheet.merged_ranges)}")
    lines.append("")
    lines.append(markdown_table(sheet.rows) if sheet.rows else "_Empty sheet._")
    if sheet.formulas:
        lines.extend(["", "## Formulas"])
        lines.extend(f"- {formula['cell']}: {formula['formula']}" for formula in sheet.formulas)
    return "\n".join(lines).strip() + "\n"


def spreadsheet_sheet_span_text(sheet: ExtractedSpreadsheetSheet, record: dict[str, str | int] | None) -> str:
    lines = [
        f"Extracted spreadsheet sheet {sheet.name}.",
        f"Range: {sheet.range_ref}. Rows: {len(sheet.rows)}. Columns: {max((len(row) for row in sheet.rows), default=0)}.",
    ]
    if record and record.get("markdown"):
        lines.append(f"Artifact: {record['markdown']}.")
    if sheet.merged_ranges:
        lines.append(f"Merged ranges: {', '.join(sheet.merged_ranges)}.")
    if sheet.formulas:
        formula_summary = "; ".join(f"{formula['cell']} {formula['formula']}" for formula in sheet.formulas[:20])
        lines.append(f"Formulas: {formula_summary}.")
    lines.append(markdown_table(sheet.rows) if sheet.rows else "Sheet is empty.")
    return "\n\n".join(lines).strip()


def spreadsheet_cell_text(formula_value: Any, cached_value: Any) -> str:
    if isinstance(formula_value, str) and formula_value.startswith("="):
        return spreadsheet_value_text(cached_value) or formula_value
    return spreadsheet_value_text(formula_value)


def spreadsheet_value_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if hasattr(value, "item"):
        try:
            value = value.item()
        except ValueError:
            pass
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    return " ".join(str(value).split())


def trim_spreadsheet_grid(rows: list[list[str]]) -> tuple[list[list[str]], str]:
    occupied = [
        (row_index, column_index)
        for row_index, row in enumerate(rows, start=1)
        for column_index, value in enumerate(row, start=1)
        if value
    ]
    if not occupied:
        return [], "A1"
    min_row = min(row for row, _ in occupied)
    max_row = max(row for row, _ in occupied)
    min_column = min(column for _, column in occupied)
    max_column = max(column for _, column in occupied)
    cropped = [row[min_column - 1:max_column] for row in rows[min_row - 1:max_row]]
    return cropped, f"{spreadsheet_cell_ref(min_row, min_column)}:{spreadsheet_cell_ref(max_row, max_column)}"


def spreadsheet_cell_ref(row: int, column: int) -> str:
    return f"{spreadsheet_column_name(column)}{row}"


def spreadsheet_column_name(column: int) -> str:
    name = ""
    while column:
        column, remainder = divmod(column - 1, 26)
        name = chr(ord("A") + remainder) + name
    return name or "A"


def schematic_artifact_from_pdf_page(
    *,
    schematic_id: str,
    filename: str,
    page_number: int,
    figure_id: str,
    image_path: str,
    image_file: Path,
    page_text: str,
    visual_summary: dict[str, int],
) -> SchematicArtifact | None:
    metrics = schematic_image_metrics(image_file)
    references = reference_designators(page_text)
    labels = schematic_labels(page_text)
    reasons = schematic_reasons(filename, page_text, references, labels, visual_summary, metrics)
    if not reasons:
        return None
    return SchematicArtifact(
        id=schematic_id,
        source="pdf-page",
        locator=f"page {page_number} {figure_id}",
        image=image_path,
        width=int(metrics["width"]),
        height=int(metrics["height"]),
        reasons=reasons,
        reference_designators=references,
        labels=labels,
        connection_cues=schematic_connection_cues(visual_summary, metrics),
        metrics=metrics,
    )


def schematic_artifact_from_image(
    *,
    schematic_id: str,
    filename: str,
    frame_index: int,
    frames: int,
    image_path: str,
    image: Image.Image,
) -> SchematicArtifact | None:
    metrics = schematic_image_metrics(image)
    references = reference_designators(filename)
    labels = schematic_labels(filename)
    locator = "image" if frames == 1 else f"image frame {frame_index}"
    reasons = schematic_reasons(filename, filename, references, labels, {}, metrics)
    if not reasons:
        return None
    return SchematicArtifact(
        id=schematic_id,
        source="image",
        locator=locator,
        image=image_path,
        width=int(metrics["width"]),
        height=int(metrics["height"]),
        reasons=reasons,
        reference_designators=references,
        labels=labels,
        connection_cues=schematic_connection_cues({}, metrics),
        metrics=metrics,
    )


def schematic_reasons(
    filename: str,
    source_text: str,
    references: list[str],
    labels: list[str],
    visual_summary: dict[str, int],
    metrics: dict[str, int | float],
) -> list[str]:
    reasons: list[str] = []
    lower_text = f"{filename}\n{source_text}".lower()
    has_schematic_terms = any(term in lower_text for term in SCHEMATIC_TERMS)
    if has_schematic_terms:
        reasons.append("source text or filename contains schematic/circuit terms")
    if len(references) >= 2:
        reasons.append("source text contains multiple reference-designator-like labels")
    if labels and (has_schematic_terms or references):
        reasons.append("source text contains schematic net or signal labels")
    if visual_summary and references and (visual_summary.get("lines", 0) + visual_summary.get("curves", 0) + visual_summary.get("rectangles", 0)) >= 6:
        reasons.append("rendered page has schematic-like line geometry near electrical labels")
    if not visual_summary and any(term in lower_text for term in {"schematic", "circuit"}) and float(metrics.get("edge_ratio", 0.0)) >= 0.01:
        reasons.append("image filename indicates a schematic and the image has line-art edges")
    return reasons


def schematic_connection_cues(visual_summary: dict[str, int], metrics: dict[str, int | float]) -> list[str]:
    cues: list[str] = []
    line_count = visual_summary.get("lines", 0)
    curve_count = visual_summary.get("curves", 0)
    rectangle_count = visual_summary.get("rectangles", 0)
    if line_count:
        cues.append(f"{line_count} PDF vector line objects")
    if curve_count:
        cues.append(f"{curve_count} PDF vector curve objects")
    if rectangle_count:
        cues.append(f"{rectangle_count} PDF rectangle objects")
    edge_ratio = float(metrics.get("edge_ratio", 0.0))
    dark_ratio = float(metrics.get("dark_ratio", 0.0))
    if edge_ratio:
        cues.append(f"line-art edge ratio {edge_ratio:.3f}")
    if dark_ratio:
        cues.append(f"dark-pixel ratio {dark_ratio:.3f}")
    return cues


def schematic_image_metrics(image_or_path: Image.Image | Path) -> dict[str, int | float]:
    if isinstance(image_or_path, Path):
        with Image.open(image_or_path) as image:
            return schematic_image_metrics(image)
    grayscale = image_or_path.convert("L")
    width, height = grayscale.size
    pixels = max(1, width * height)
    dark_pixels = sum(grayscale.histogram()[:96])
    edges = grayscale.filter(ImageFilter.FIND_EDGES)
    edge_pixels = sum(edges.histogram()[33:])
    return {
        "width": width,
        "height": height,
        "dark_ratio": round(dark_pixels / pixels, 6),
        "edge_ratio": round(edge_pixels / pixels, 6),
    }


def reference_designators(text: str) -> list[str]:
    return sorted(set(REFERENCE_DESIGNATOR_RE.findall(text.upper())))


def schematic_labels(text: str) -> list[str]:
    return sorted(set(NET_LABEL_RE.findall(text.upper().replace("_", " ").replace("-", " "))))


def write_schematic_artifact(artifact_root: Path, schematic: SchematicArtifact) -> dict[str, Any]:
    schematic_dir = artifact_root / "schematics" / schematic.id
    schematic_dir.mkdir(parents=True, exist_ok=True)
    description_path = schematic_dir / "description.md"
    json_path = schematic_dir / "analysis.json"
    record: dict[str, Any] = {
        "id": schematic.id,
        "schema_version": SCHEMATIC_SCHEMA_VERSION,
        "source": schematic.source,
        "locator": schematic.locator,
        "image": schematic.image,
        "description": description_path.relative_to(artifact_root).as_posix(),
        "json": json_path.relative_to(artifact_root).as_posix(),
        "references": ", ".join(schematic.reference_designators),
        "labels": ", ".join(schematic.labels),
        "connection_cues": "; ".join(schematic.connection_cues),
        "reasons": "; ".join(schematic.reasons),
        "analysis_outputs": "[]",
    }
    description_path.write_text(schematic_markdown_document(schematic), encoding="utf-8")
    json_path.write_text(json.dumps(schematic_json_document(schematic), indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return record


def write_schematic_index(artifact_root: Path, records: list[dict[str, Any]]) -> None:
    write_tsv(
        artifact_root / "schematic_index.tsv",
        records,
        ["id", "schema_version", "source", "locator", "image", "description", "json", "references", "labels", "connection_cues", "reasons", "analysis_outputs"],
    )


def schematic_json_document(schematic: SchematicArtifact) -> dict[str, Any]:
    return {
        "schemaVersion": SCHEMATIC_SCHEMA_VERSION,
        "id": schematic.id,
        "source": schematic.source,
        "locator": schematic.locator,
        "image": schematic.image,
        "imageSize": {"width": schematic.width, "height": schematic.height},
        "classification": {"isSchematic": True, "reasons": schematic.reasons, "metrics": schematic.metrics},
        "referenceDesignators": schematic.reference_designators,
        "labels": schematic.labels,
        "connectionCues": schematic.connection_cues,
        "analysisOutputs": [],
    }


def schematic_markdown_document(schematic: SchematicArtifact) -> str:
    lines = [
        f"# Schematic Artifact {schematic.id}",
        "",
        f"Source locator: {schematic.locator}",
        f"Image artifact: {schematic.image}",
        f"Schema version: {SCHEMATIC_SCHEMA_VERSION}",
        "",
        "## Classification",
        *[f"- {reason}" for reason in schematic.reasons],
        "",
        "## Visible Text Candidates",
        f"Reference designators: {', '.join(schematic.reference_designators) if schematic.reference_designators else 'not extracted by the deterministic Phase 1 analyzer'}",
        f"Labels and nets: {', '.join(schematic.labels) if schematic.labels else 'not extracted by the deterministic Phase 1 analyzer'}",
        "",
        "## Connection Cues",
        *[f"- {cue}" for cue in schematic.connection_cues],
        "",
        "## Structured Analysis Outputs",
        "No component detection, connectivity mapping, OCR assignment, or netlist output is attached yet. Future SINA-style analyzers can append outputs to this schema without re-ingesting the source document.",
    ]
    return "\n".join(lines).strip() + "\n"


def schematic_span_text(schematic: SchematicArtifact, record: dict[str, Any]) -> str:
    references = ", ".join(schematic.reference_designators) if schematic.reference_designators else "not extracted"
    labels = ", ".join(schematic.labels) if schematic.labels else "not extracted"
    cues = "; ".join(schematic.connection_cues) if schematic.connection_cues else "no geometric cues recorded"
    reasons = "; ".join(schematic.reasons)
    return "\n".join([
        f"Schematic image artifact {schematic.id} from {schematic.locator}.",
        f"Description artifact: {record['description']}. Image artifact: {schematic.image}. JSON artifact: {record['json']}.",
        f"Classification reasons: {reasons}.",
        f"Reference designator candidates: {references}. Labels and net candidates: {labels}.",
        f"Connection cues: {cues}.",
        "Analysis outputs are empty in Phase 1; this record is schema-ready for future component detections, connectivity mappings, OCR/designator assignments, and SPICE netlists.",
    ]).strip()


def render_page_artifact(pdf: pdfium.PdfDocument, page_number: int, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    page = pdf[page_number - 1]
    try:
        page.render(scale=2).to_pil().save(output_path)
    finally:
        page.close()
    return output_path


def page_visual_summary(page: pdfplumber.page.Page) -> dict[str, int] | None:
    summary = {
        "images": len(page.images),
        "lines": len(page.lines),
        "rectangles": len(page.rects),
        "curves": len(page.curves),
    }
    return summary if any(summary.values()) else None


def write_index_files(artifact_root: Path, tables: list[dict[str, str | int]], figures: list[dict[str, str | int]], schematics: list[dict[str, Any]] | None = None) -> None:
    if tables:
        write_tsv(artifact_root / "table_index.tsv", tables, ["id", "page", "rows", "columns", "non_empty_cells", "total_cells", "csv", "markdown"])
    if figures:
        write_tsv(artifact_root / "figure_index.tsv", figures, ["id", "page", "kind", "path", "images", "lines", "rectangles", "curves"])
        lines = ["# Extracted Visual Artifacts", ""]
        for figure in figures:
            lines.append(
                f"- {figure['id']} page {figure['page']}: {figure['path']} "
                f"(images={figure['images']}, lines={figure['lines']}, rectangles={figure['rectangles']}, curves={figure['curves']})"
            )
        (artifact_root / "figures.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    if schematics:
        write_schematic_index(artifact_root, schematics)


def write_tsv(path: Path, rows: list[dict[str, str | int]], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def table_span_text(table_id: str, page_number: int, table: ExtractedTable) -> str:
    return "\n".join([
        f"Extracted PDF table {table_id} from page {page_number}.",
        "Review the page render for layout-sensitive or visually dense tables.",
        table_markdown_document(table),
    ]).strip()


def visual_span_text(figure_id: str, page_number: int, summary: dict[str, int]) -> str:
    return (
        f"Extracted PDF visual artifact {figure_id} from page {page_number}. "
        f"Rendered page image preserves graphs, plots, diagrams, flowcharts, schematics, and other visual layout. "
        f"Detected objects: images={summary['images']}, lines={summary['lines']}, "
        f"rectangles={summary['rectangles']}, curves={summary['curves']}."
    )


def image_span_text(metadata: dict[str, Any]) -> str:
    artifact = f" Artifact: {metadata['artifact']}." if "artifact" in metadata else ""
    artifacts = metadata.get("artifacts")
    frame_artifacts = f" Frame artifacts: {', '.join(artifacts)}." if isinstance(artifacts, list) and len(artifacts) > 1 else ""
    return (
        f"Image source {metadata['filename']}. Format {metadata['format']}; "
        f"size {metadata['width']}x{metadata['height']}; mode {metadata['mode']}; frames {metadata['frames']}."
        f"{artifact}{frame_artifacts} Use these visual artifacts for diagrams, graphs, flowcharts, screenshots, scanned pages, and photos."
    )


def markdown_table(table: list[list[str]]) -> str:
    width = max((len(row) for row in table), default=0)
    if width == 0:
        return ""
    rows = [row + [""] * (width - len(row)) for row in table]
    header = rows[0]
    body = rows[1:] or [[""] * width]
    lines = [
        "| " + " | ".join(escape_table_cell(cell) for cell in header) + " |",
        "| " + " | ".join("---" for _ in range(width)) + " |",
        *["| " + " | ".join(escape_table_cell(cell) for cell in row) + " |" for row in body],
    ]
    return "\n".join(lines)


def table_markdown_document(table: ExtractedTable) -> str:
    summary = register_bit_summary(table.rows)
    if summary:
        return f"{summary}\n\n{markdown_table(table.rows)}"
    return markdown_table(table.rows)


def register_bit_summary(rows: list[list[str]]) -> str:
    if not rows or not is_register_bit_header(rows[0]):
        return ""
    width = len(rows[0])
    bit_headers = [int(cell, 10) for cell in rows[0][1:]]
    values_by_label = {
        normalized_table_label(row[0]): (row + [""] * (width - len(row)))[1:width]
        for row in rows[1:]
    }
    field_values = values_by_label.get("field")
    if not field_values:
        return ""
    labels = [
        ("Field", field_values),
        ("Reset", values_by_label.get("reset", [""] * len(bit_headers))),
        ("Access Type", values_by_label.get("access type", values_by_label.get("access", [""] * len(bit_headers)))),
    ]
    lines = ["Register bit layout:"]
    start = 0
    while start < len(bit_headers):
        current = tuple(values[start] for _, values in labels)
        end = start + 1
        while end < len(bit_headers) and tuple(values[end] for _, values in labels) == current:
            end += 1
        description = "; ".join(f"{label} {value}" for label, values in labels if (value := values[start]))
        if description:
            lines.append(f"- {format_bit_range(bit_headers[start:end])}: {description}")
        start = end
    return "\n".join(lines) if len(lines) > 1 else ""


def format_bit_range(bits: list[int]) -> str:
    if len(bits) == 1:
        return f"bit {bits[0]}"
    return f"bits {bits[0]}:{bits[-1]}"


def escape_table_cell(value: str) -> str:
    return " ".join(value.replace("|", "\\|").split())


def frame_count(image: Image.Image) -> int:
    try:
        return sum(1 for _ in ImageSequence.Iterator(image))
    except Exception:
        return 1


def extract_html(content: bytes) -> str:
    soup = BeautifulSoup(content, "html.parser")
    for element in soup(["script", "style", "template", "noscript"]):
        element.extract()
    return "\n".join(line.strip() for line in soup.get_text("\n").splitlines() if line.strip())


def decode_text(content: bytes) -> str:
    return content.decode("utf-8", errors="replace").strip()


def source_suffix(name: str) -> str:
    return Path(urlparse(name).path or name).suffix.lower()


def normalized_content_type(content_type: str | None) -> str:
    return (content_type or "").split(";", 1)[0].strip().lower()


def looks_like_pdf(content: bytes) -> bool:
    return content.lstrip()[:5] == b"%PDF-"


def safe_stem(name: str) -> str:
    stem = Path(name).stem or "image"
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", stem.strip())[:80].strip("._-")
    return safe or "image"
